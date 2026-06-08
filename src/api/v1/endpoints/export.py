import openpyxl

from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
from io import BytesIO

from typing import Any, List

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from api import deps  # noqa

import crud, models, schemas  # noqa
from services.export.report_rows import build_report_rows

router = APIRouter()


@router.get('/proxies')
async def export_proxies(
    *,
    db: AsyncSession = Depends(deps.get_db),
    filters: List[schemas.Filter] = Depends(deps.request_filters),
    current_user: models.User = Depends(deps.get_current_active_user)
):
    if crud.user.is_superuser(current_user):
        proxies = await crud.proxy.get_all(db=db, filters=filters)
    else:
        proxies = await crud.proxy.get_all_by_user(db=db, filters=filters)

    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Proxies'

    headers = [
        '№', 'Name', 'URL', 'APIKey', 'Good', 'Bad',
        'Last used', 'Last used successful'
    ]
    sheet.append(headers)

    for proxy in proxies:
        sheet.append([
            proxy.id,
            proxy.name,
            proxy.url,
            proxy.good_count,
            proxy.bad_count,
            proxy.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                if proxy.timestamp else '',
            proxy.ts_1.strftime('%Y-%m-%d %H:%M:%S')
            if proxy.ts_1 else '',
        ])

    table_ref = 'A1:{}'.format(
        sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate
    )
    table = Table(displayName='DataTable', ref=table_ref)
    style = TableStyleInfo(
        name='TableStyleMedium9',
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    for column in sheet.columns:
        adjusted_width = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[
            column[0].column_letter].width = adjusted_width + 5

    workbook.save(output)
    output.seek(0)

    filename = f'o3go_stats_proxies_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}.xlsx'
    headers = {
        'Content-Disposition': f'attachment; filename={filename}'
    }

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-'
                   'officedocument.spreadsheetml.sheet',
        headers=headers
    )


@router.get('/numbers')
async def export_numbers(
    *,
    db: AsyncSession = Depends(deps.get_db),
    filters: List[schemas.Filter] = Depends(deps.request_filters),
    current_user: models.User = Depends(deps.get_current_active_user)
):
    if crud.user.is_superuser(current_user):
        numbers = await crud.number.get_all(db=db, filters=filters)
    else:
        numbers = await crud.number.get_all_by_user(db=db, filters=filters)

    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Numbers'

    headers = [
        '№', 'Number', 'Service', 'APIKey', 'Device ID', 'Proxy',
        'Timestamp', 'Info_1', 'Info_2', 'Info_3'
    ]
    sheet.append(headers)

    for number in numbers:
        sheet.append([
            number.id,
            number.number,
            number.api_key,
            number.service_alias,
            number.device_ext_id,
            number.proxy,
            number.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                if number.timestamp else '',
            number.info_1,
            number.info_2,
            number.info_3
        ])

    table_ref = 'A1:{}'.format(
        sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate
    )
    table = Table(displayName='DataTable', ref=table_ref)
    style = TableStyleInfo(
        name='TableStyleMedium9',
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    for column in sheet.columns:
        adjusted_width = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[
            column[0].column_letter].width = adjusted_width + 5

    workbook.save(output)
    output.seek(0)

    filename = f'o3go_stats_numbers_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}.xlsx'
    headers = {
        'Content-Disposition': f'attachment; filename={filename}'
    }

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-'
                   'officedocument.spreadsheetml.sheet',
        headers=headers
    )


@router.get('/report')
async def export_report(
    *,
    db: AsyncSession = Depends(deps.get_db),
    filters: List[schemas.Filter] = Depends(deps.request_filters),
    current_user: models.User = Depends(deps.get_current_active_user)
):
    # Множество счётчиков — в репорте имеют суффикс _count (MyService_code_count)
    COUNT_SUFFIX = {
        'start', 'number', 'code', 'no_code', 'waiting', 'bad',
        'error_1', 'error_2', 'account', 'account_ban', 'sent', 'delivered',
    }

    def expand_column(col: str, svc_key: str) -> tuple[str | None, str]:
        # Преобразует короткое имя колонки из Service.columns в полный ключ репорта.
        # Возвращает (ключ_репорта, короткая_ярлык). Если ключ None — колонка вычисляемая.
        if col in ('code_pct', 'sent_avg'):
            # Вычисляемые колонки без прямого ключа в репорте
            return (None, col)
        if col in COUNT_SUFFIX:
            return (f'{svc_key}_{col}_count', col)
        # Стоимости и всего — без суффикса _count
        return (f'{svc_key}_{col}', col)

    # Базовые колонки берутся напрямую из данных репорта (ключи = имена полей)
    col_keys = [
        ('api_key', None, 'api_key'),
        ('device_id', None, 'device_id'),
        ('device_ext_id', None, 'device_ext_id'),
        ('device_root', None, 'device_root'),
        ('device_operator', None, 'device_operator'),
    ]
    # Заголовки для базовых колонок
    excel_headers = ['API Key', 'Device', 'Ext ID', 'Root', 'Operator']

    reports, services = await build_report_rows(
        db=db, filters=filters, current_user=current_user
    )

    # Формируем сервисные колонки из Service.columns каждого сервиса
    # Ключ сервиса для репорта должен совпадать с ключом из crud.report.get_all():
    # (svc.name or svc.alias.title()).replace(' ', '_')
    for svc in services:
        # Ключ: Telegram, Viber, WhatsApp и т.д.
        svc_key = (svc.name or svc.alias.title()).replace(' ', '_')
        # Отображаемое имя: name службы или title(alias)
        display_name = svc.name or svc.alias.title()

        if not svc.columns:
            # Если для сервиса не настроены колонки — пропускаем
            continue

        for col in svc.columns:
            report_key, short_label = expand_column(col, svc_key)
            # Заголовок: "{ServiceName} {ColumnLabel}"
            excel_headers.append(f'{display_name} {short_label.replace("_", " ").title()}')
            # col_keys хранит кортеж (report_key, svc_key, short_label):
            # - report_key не None: обычные данные из репорта (например "WhatsApp_code_count")
            # - report_key None: вычисляемая колонка (например code_pct)
            col_keys.append((report_key, svc_key, short_label))

    # Итоговые стоимости — всегда показываем
    col_keys.extend([
        ('code_total', None, 'code_total'),
        ('sent_total', None, 'sent_total'),
    ])
    excel_headers.extend(['Total Code Cost', 'Total Sent Cost'])

    # Глобальные колонки — всегда показываем
    col_keys.extend([
        ('timestamp', None, 'timestamp'),
        ('ts_1', None, 'ts_1'),
        ('info_1', None, 'info_1'),
        ('info_2', None, 'info_2'),
        ('info_3', None, 'info_3'),
    ])
    excel_headers.extend(['Last Activity', 'Last Success Code', 'Info 1', 'Info 2', 'Info 3'])

    for report in reports:
        # Форматируем даты в строки для Excel
        if report.get('timestamp'):
            report['timestamp'] = report['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
        if report.get('ts_1'):
            report['ts_1'] = report['ts_1'].strftime('%Y-%m-%d %H:%M:%S')

    # Функция получения значения ячейки по спецификации колонки.
    # spec = (report_key, svc_key, short_label) — элементы col_keys.
    def get_cell(report, key, svc_key, short_label):
        # Обычная колонка — берём значение из отчёта
        if key:
            return report.get(key, 0)
        # code_pct: (code_count / start_count) * 100
        if short_label == 'code_pct':
            start = report.get(f'{svc_key}_start_count', 0) or 0
            code = report.get(f'{svc_key}_code_count', 0) or 0
            return round(code / start * 100, 2) if start else 0
        # sent_avg: (delivered_count / sent_count) * 100
        if short_label == 'sent_avg':
            sent = report.get(f'{svc_key}_sent_count', 0) or 0
            deliv = report.get(f'{svc_key}_delivered_count', 0) or 0
            return round(deliv / sent * 100, 2) if sent else 0
        return ''

    # --- Запись в Excel ---
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Report'
    sheet.append(excel_headers)

    # Для каждого отчёта формируем строку значений по всем заданным колонкам
    for report in reports:
        sheet.append([get_cell(report, *spec) for spec in col_keys])

    # Добавляем форматированную таблицу
    table_ref = 'A1:{}'.format(
        sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate
    )
    table = Table(displayName='DataTable', ref=table_ref)
    style = TableStyleInfo(
        name='TableStyleMedium9',
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Автоподбор ширины столбцов
    for column in sheet.columns:
        adjusted_width = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width + 5

    workbook.save(output)
    output.seek(0)

    filename = f'o3go_stats_report_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}.xlsx'
    http_headers = {
        'Content-Disposition': f'attachment; filename={filename}'
    }

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=http_headers
    )
