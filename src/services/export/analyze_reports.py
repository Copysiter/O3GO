#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Анализатор отчётов по сотрудникам.

Использование:
    python analyze_reports.py [папка_с_xlsx] [папка_для_отчёта]

По умолчанию:
    папка_с_xlsx     = текущая директория
    папка_для_отчёта = текущая директория

Файлы в папке должны быть .xlsx, имя файла = имя сотрудника.
В каждом файле — лист с колонками формата (API Key, Ext ID, Operator,
Click/Wa/Business/Regular/Viber/Kakao Code/Sent ..., Total Code Cost,
Total Sent Cost, Timestamp, Timedelta, ...).

Файл может содержать итоговую строку (с пустым API Key) — она игнорируется.
Если в файле несколько разных API Key, скрипт автоматически детектит
"сборные" файлы (>=3 разных API Key и название содержит подсказку
вроде "сборн", "неучт", "разное") и разбивает их на под-сотрудников.

Результат:
    <папка_для_отчёта>/Отчёт_по_сотрудникам.xlsx
    <папка_для_отчёта>/Отчёт_по_сотрудникам.html
"""

import sys
import os
import html
import json
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
except ImportError:
    print("ОШИБКА: не установлен pandas. Установи: pip install pandas openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
except ImportError:
    print("ОШИБКА: не установлен openpyxl. Установи: pip install openpyxl")
    sys.exit(1)


# =============================================================
# 1. ЗАГРУЗКА И ПОДГОТОВКА ДАННЫХ
# =============================================================

REQUIRED_COLS = ['API Key', 'Total Code Cost', 'Total Sent Cost']

# =============================================================
# ЦЕЛЕВОЙ ПОКАЗАТЕЛЬ
# =============================================================
# Цель: каждое устройство должно приносить TARGET_PER_DEVICE_DAY рублей в сутки.
# Если бизнес-цель изменится — меняйте это значение, всё в отчёте пересчитается.
TARGET_PER_DEVICE_DAY = 50.0  # ₽/устр/сутки

# Признаки "сборного" файла — если они есть в названии, и API Key > 1,
# то файл разбивается на под-сотрудников
SPLIT_HINTS = ['сборн', 'неучт', 'свалк', 'разное', 'разн', 'mixed', 'misc']


def looks_like_split_file(filename_stem: str, n_api_keys: int) -> bool:
    """Решаем, надо ли разбивать сборный файл на под-сотрудников."""
    if n_api_keys < 2:
        return False
    name_lower = filename_stem.lower()
    if any(h in name_lower for h in SPLIT_HINTS):
        return True
    # Если в названии 3+ слова через _ или пробел, и API Key тоже >=3 — почти всегда сборник
    parts = [p for p in filename_stem.replace('_', ' ').split() if p]
    if len(parts) >= 3 and n_api_keys >= 3:
        return True
    return False


# Старая функция load_all_files заменена новой ниже —
# она объединяет оба режима: по конфигу и по имени файла.
# Логика старого режима теперь в _load_by_filename().



# =============================================================
# КОНФИГ-БАЗА: МАТЧИНГ API KEY → СОТРУДНИК
# =============================================================

import re


def normalize_key(k: str) -> str:
    """Нормализация API Key для сравнения: без пробелов, верхний регистр."""
    if k is None:
        return ''
    return re.sub(r'\s+', '', str(k).strip().upper())


def expand_pattern(pattern: str) -> list:
    """
    Раскрывает шаблоны вида SPB[01-20] в список ['SPB01','SPB02',...,'SPB20'].
    Если паттерн без скобок — возвращает [pattern] как есть.
    Поддерживает форматы [01-20] (с ведущим нулём) и [1-20] (без).
    """
    m = re.match(r'^(.*?)\[(\d+)-(\d+)\](.*)$', pattern)
    if not m:
        return [pattern]
    prefix, start_str, end_str, suffix = m.groups()
    start, end = int(start_str), int(end_str)
    # Сохраняем ведущие нули: если start_str = '01' → ширина 2
    width = len(start_str) if start_str.startswith('0') and len(start_str) > 1 else 0
    result = []
    for i in range(start, end + 1):
        num = f'{i:0{width}d}' if width else str(i)
        result.append(f'{prefix}{num}{suffix}')
    return result


def load_employee_config(config_path: Path) -> dict:
    """
    Читает YAML-конфиг с маппингом сотрудник → [API Key].
    Возвращает dict { normalized_key: employee_name }.
    Также возвращает обратный маппинг { employee_name: [keys] } для отчёта.
    """
    if not config_path.exists():
        return None

    # Парсим YAML вручную (без зависимости от pyyaml)
    config = parse_simple_yaml(config_path)

    if 'employees' not in config:
        raise SystemExit(f"В конфиге {config_path} нет ключа 'employees'")

    employees = config['employees']
    key_to_emp = {}      # normalized_key → employee_name
    emp_to_keys = {}     # employee_name → [original_keys]
    conflicts = []

    for emp, keys in employees.items():
        if not isinstance(keys, list):
            continue
        expanded = []
        for k in keys:
            expanded.extend(expand_pattern(str(k)))
        emp_to_keys[emp] = expanded
        for k in expanded:
            nk = normalize_key(k)
            if nk in key_to_emp and key_to_emp[nk] != emp:
                conflicts.append((nk, key_to_emp[nk], emp))
            key_to_emp[nk] = emp

    if conflicts:
        print("⚠ КОНФЛИКТЫ В КОНФИГЕ: один API Key привязан к нескольким сотрудникам:")
        for nk, e1, e2 in conflicts:
            print(f"  {nk}: {e1} ↔ {e2}")
        print("Использую последнее назначение. Поправь конфиг!")

    return {
        'key_to_emp': key_to_emp,
        'emp_to_keys': emp_to_keys,
        'all_employees': list(employees.keys()),
    }


def parse_simple_yaml(path: Path) -> dict:
    """
    Минимальный YAML-парсер только для нужного нам формата:
      employees:
        Имя:
          - KEY1
          - KEY2
    Чтобы не тащить pyyaml как зависимость.
    """
    result = {'employees': {}}
    current_emp = None

    with open(path, 'r', encoding='utf-8') as f:
        for raw_line in f:
            # Удаляем комментарии (но не внутри строк — у нас простой формат, можно так)
            line = raw_line.split('#')[0].rstrip()
            if not line.strip():
                continue

            # Уровень отступа
            stripped = line.lstrip()
            indent = len(line) - len(stripped)

            if stripped == 'employees:':
                continue
            elif stripped.startswith('- '):
                # Это API Key
                key = stripped[2:].strip()
                if current_emp:
                    result['employees'][current_emp].append(key)
            elif stripped.endswith(':') and indent >= 2:
                # Имя сотрудника
                current_emp = stripped[:-1].strip()
                result['employees'][current_emp] = []

    return result


def load_with_config(files: list, config: dict) -> tuple:
    """
    Загружает данные и разбивает их по API Key согласно конфигу.

    Возвращает:
      all_data — DataFrame со столбцом '_Сотрудник'
      employees_order — список сотрудников (в порядке появления в конфиге)
      unmatched_keys — список API Key из данных, не найденных в конфиге
    """
    all_rows = []
    skipped = []

    for fp in files:
        try:
            df = pd.read_excel(fp)
        except Exception as e:
            skipped.append(f"{fp.name}: {e}")
            continue

        if not all(c in df.columns for c in REQUIRED_COLS):
            skipped.append(f"{fp.name}: нет нужных колонок")
            continue

        df = df[df['API Key'].notna()].copy()
        if df.empty:
            continue

        # Матчим каждую строку с сотрудником через нормализованный ключ
        df['_norm_key'] = df['API Key'].apply(normalize_key)
        df['_Сотрудник'] = df['_norm_key'].map(config['key_to_emp']).fillna('Неизвестно')
        all_rows.append(df)

    if not all_rows:
        raise SystemExit("Не удалось загрузить ни одного валидного файла")

    all_data = pd.concat(all_rows, ignore_index=True)

    # Находим неизвестные API Key
    unmatched = sorted(
        all_data[all_data['_Сотрудник'] == 'Неизвестно']['API Key'].unique().tolist()
    )

    # Порядок сотрудников: сначала из конфига, потом 'Неизвестно' если есть
    present_employees = all_data['_Сотрудник'].unique().tolist()
    employees_order = [e for e in config['all_employees'] if e in present_employees]
    if 'Неизвестно' in present_employees:
        employees_order.append('Неизвестно')

    if skipped:
        print("Пропущены файлы:")
        for s in skipped:
            print(f"  - {s}")

    return all_data, employees_order, unmatched


def load_all_files(folder: Path, config_path: Path = None):
    """
    Универсальный загрузчик с двумя режимами:

    1. Если есть config_path (/storage/employees.yml) — используем матчинг по API Key.
       Подходит для "цельного файла из системы".

    2. Если config_path нет — старая логика: имя файла = сотрудник,
       сборные файлы разбиваются по подсказкам в имени.

    Возвращает (all_data, employees_order, mode_info)
      mode_info — dict с полями {'mode': 'config'|'filename', 'unmatched_keys': [...]}
    """
    files = sorted([p for p in folder.glob('*.xlsx') if not p.name.startswith('~$')])
    if not files:
        raise SystemExit(f"В папке {folder} не найдено .xlsx файлов")

    # Если задан конфиг — используем его
    if config_path and config_path.exists():
        config = load_employee_config(config_path)
        if config:
            print(f"  Режим: матчинг по конфигу {config_path.name}")
            print(f"  Сотрудников в конфиге: {len(config['all_employees'])}")
            all_data, emp_order, unmatched = load_with_config(files, config)
            return all_data, emp_order, {
                'mode': 'config',
                'unmatched_keys': unmatched,
                'config': config,
            }

    # Иначе старый режим
    print("  Режим: имя файла = сотрудник (конфиг не найден)")
    all_data, emp_order = _load_by_filename(files)
    return all_data, emp_order, {'mode': 'filename', 'unmatched_keys': [], 'config': None}


def _load_by_filename(files: list) -> tuple:
    """Старая логика: имя файла = сотрудник."""
    all_rows = []
    employees_order = []
    skipped = []

    for fp in files:
        try:
            df = pd.read_excel(fp)
        except Exception as e:
            skipped.append(f"{fp.name}: {e}")
            continue

        if not all(c in df.columns for c in REQUIRED_COLS):
            skipped.append(f"{fp.name}: нет нужных колонок")
            continue

        df = df[df['API Key'].notna()].copy()
        if df.empty:
            continue

        employee_name = fp.stem
        n_keys = df['API Key'].nunique()

        if looks_like_split_file(employee_name, n_keys):
            for api_key, sub in df.groupby('API Key'):
                sub = sub.copy()
                sub_name = f"{employee_name}_{api_key}"
                sub['_Сотрудник'] = sub_name
                all_rows.append(sub)
                if sub_name not in employees_order:
                    employees_order.append(sub_name)
        else:
            df['_Сотрудник'] = employee_name
            all_rows.append(df)
            if employee_name not in employees_order:
                employees_order.append(employee_name)

    if not all_rows:
        raise SystemExit("Не удалось загрузить ни одного валидного файла")

    all_data = pd.concat(all_rows, ignore_index=True)

    if skipped:
        print("Пропущены файлы:")
        for s in skipped:
            print(f"  - {s}")

    return all_data, employees_order


def enrich(all_data: pd.DataFrame) -> pd.DataFrame:
    """Добавляем расчётные колонки. Все недостающие колонки заполняются нулями."""

    # Безопасное обращение к колонке (если её нет — 0)
    def col(name, default=0):
        if name in all_data.columns:
            return pd.to_numeric(all_data[name], errors='coerce').fillna(default)
        return pd.Series([default] * len(all_data), index=all_data.index)

    all_data['Итого_руб'] = col('Total Code Cost') + col('Total Sent Cost')

    # Доходы по каналам
    all_data['Доход_Click'] = col('click Code Total')
    all_data['Доход_Wa'] = col('wa_load Sent Total')
    all_data['Доход_Business'] = col('business Code Total')
    all_data['Доход_Regular'] = col('regular Code Total') + col('regular Sent Total')
    all_data['Доход_Viber'] = col('viber Code Total') + col('viber Sent Total')
    all_data['Доход_Kakao'] = col('kakao Code Total') + col('kakao Sent Total')

    # Метрики качества и активности
    all_data['Коды_всего'] = (
        col('Click Code Count') + col('Business Code Count') +
        col('Regular Code Count') + col('Viber Code Count') + col('Kakao Code Count')
    )
    all_data['Bad_clicks'] = col('Click Bad Count')
    all_data['Click_total'] = col('Click Start Count')
    all_data['Wa_bans'] = col('Wa Load Account Ban Count')
    all_data['Wa_accounts'] = col('Wa Load Account Count')
    all_data['Wa_sent'] = col('Wa Load Sent Count')
    all_data['Wa_delivered'] = col('Wa Load Delivered Count')
    all_data['Wa_start'] = col('Wa Load Start Count')
    all_data['Regular_sent'] = col('Regular Sent Count')
    all_data['Regular_delivered'] = col('Regular Delivered Count')
    all_data['Regular_codes'] = col('Regular Code Count')
    all_data['Regular_start'] = col('Regular Start Count')
    all_data['Viber_start'] = col('Viber Start Count')
    all_data['Kakao_start'] = col('Kakao Start Count')
    all_data['Kakao_codes'] = col('Kakao Code Count')

    # Timedelta в часах (предполагаем именно часы)
    if 'Timedelta' in all_data.columns:
        all_data['Timedelta_h'] = pd.to_numeric(all_data['Timedelta'], errors='coerce')
    else:
        all_data['Timedelta_h'] = float('nan')

    all_data['Активен_24ч'] = (all_data['Timedelta_h'] < 24).fillna(False).astype(int)
    all_data['Нулевой_доход'] = (all_data['Итого_руб'] == 0).astype(int)
    all_data['Мертвый'] = ((all_data['Итого_руб'] == 0) & (all_data['Timedelta_h'] > 48)).astype(int)

    return all_data


# =============================================================
# 2. АГРЕГАЦИИ
# =============================================================

def safe_div(num, denom, decimals=2):
    """Деление num/denom, при нулевом знаменателе возвращает 0."""
    num = pd.to_numeric(num, errors='coerce')
    denom = pd.to_numeric(denom, errors='coerce')
    if isinstance(num, pd.Series) and not isinstance(denom, pd.Series):
        denom = pd.Series(denom, index=num.index)
    return (num / denom.where(denom.ne(0))).fillna(0).round(decimals)


def safe_pct(num, denom, decimals=1):
    """Процент num/denom, не падает на нулях."""
    return safe_div(pd.to_numeric(num, errors='coerce') * 100, denom, decimals)


def company_goal_totals(summary: pd.DataFrame) -> tuple:
    """Итоги цели по компании с взвешиванием периода по устройствам."""
    total_devices = summary['Устройств'].sum()
    total_revenue = summary['Итого_руб'].sum()
    total_device_days = (summary['Устройств'] * summary['Период_суток']).sum()
    weighted_period_days = total_device_days / total_devices if total_devices else 0
    company_fact_day = total_revenue / total_device_days if total_device_days else 0
    return total_devices, total_revenue, total_device_days, weighted_period_days, company_fact_day


def compute_summary(all_data: pd.DataFrame, period_days: float = None) -> pd.DataFrame:
    """Сводка по сотрудникам — для рейтинга, качества и утилизации."""
    g = all_data.groupby('_Сотрудник')

    agg = pd.DataFrame({
        'Сотрудник': g.size().index,
        'Записей': g.size().values,
    }).reset_index(drop=True)

    # Используем nunique для Ext ID если он есть, иначе count
    has_ext = 'Ext ID' in all_data.columns
    if has_ext:
        agg['Устройств'] = g['Ext ID'].nunique().values
    else:
        agg['Устройств'] = g.size().values

    agg['API_Keys'] = g['API Key'].nunique().values

    metric_sums = [
        'Итого_руб', 'Доход_Click', 'Доход_Wa', 'Доход_Business',
        'Доход_Regular', 'Доход_Viber', 'Доход_Kakao',
        'Коды_всего', 'Bad_clicks', 'Click_total',
        'Wa_bans', 'Wa_accounts', 'Wa_sent', 'Wa_delivered', 'Wa_start',
        'Regular_sent', 'Regular_delivered', 'Regular_codes', 'Regular_start',
        'Viber_start', 'Kakao_codes', 'Kakao_start',
        'Активен_24ч', 'Нулевой_доход', 'Мертвый',
    ]
    sums = g[metric_sums].sum().reset_index(drop=True)
    for c in metric_sums:
        agg[c] = sums[c].values

    agg = agg.sort_values('Итого_руб', ascending=False).reset_index(drop=True)

    # Для UI-аналитики используем выбранный период. Timestamp — только fallback,
    # потому что это последняя активность, а не граница фильтра.
    if period_days is not None:
        agg['Период_часов'] = period_days * 24
        agg['Период_суток'] = period_days
    elif 'Timestamp' in all_data.columns:
        ts = pd.to_datetime(all_data['Timestamp'], errors='coerce')
        all_data['_ts_norm'] = ts
        ts_per_emp = all_data.groupby('_Сотрудник').agg(
            _ts_min=('_ts_norm', 'min'),
            _ts_max=('_ts_norm', 'max'),
        ).reset_index().rename(columns={'_Сотрудник': 'Сотрудник'})
        ts_per_emp['Период_часов'] = (
            (ts_per_emp['_ts_max'] - ts_per_emp['_ts_min']).dt.total_seconds() / 3600
        )
        # Гарантируем минимум 1 час, чтобы не делить на 0 для записей с одним timestamp
        ts_per_emp['Период_часов'] = ts_per_emp['Период_часов'].clip(lower=1.0)
        ts_per_emp['Период_суток'] = ts_per_emp['Период_часов'] / 24
        agg = agg.merge(
            ts_per_emp[['Сотрудник', 'Период_часов', 'Период_суток']],
            on='Сотрудник', how='left'
        )
    else:
        agg['Период_часов'] = 24.0
        agg['Период_суток'] = 1.0

    # Вычисляемые метрики
    agg['Доход_на_устр'] = safe_div(agg['Итого_руб'], agg['Устройств'], 2)

    # === ЦЕЛЕВЫЕ МЕТРИКИ (50 ₽/устр/сутки) ===
    # Фактический доход на устройство в сутки
    denom = agg['Устройств'] * agg['Период_суток']
    agg['Факт_руб_устр_сутки'] = safe_div(agg['Итого_руб'], denom, 2)

    # План за период: цель × устройств × дней
    agg['План_за_период'] = (
        TARGET_PER_DEVICE_DAY * agg['Устройств'] * agg['Период_суток']
    ).round(2)

    # Абсолютное отклонение от плана за период
    agg['Отклонение_за_период'] = (agg['Итого_руб'] - agg['План_за_период']).round(2)

    # Процент выполнения плана
    agg['Выполнение_pct'] = safe_pct(agg['Итого_руб'], agg['План_за_период'], 1)

    # Отклонение в ₽/устр/сутки (положительное = выше цели)
    agg['Отклонение_сутки'] = (agg['Факт_руб_устр_сутки'] - TARGET_PER_DEVICE_DAY).round(2)

    agg['Click_bad_pct'] = safe_pct(agg['Bad_clicks'], agg['Click_total'])
    agg['WA_ban_pct'] = safe_pct(agg['Wa_bans'], agg['Wa_accounts'])
    agg['WA_delivery_pct'] = safe_pct(agg['Wa_delivered'], agg['Wa_sent'])
    agg['Reg_delivery_pct'] = safe_pct(agg['Regular_delivered'], agg['Regular_sent'])
    agg['Reg_codes_pct'] = safe_pct(agg['Regular_codes'], agg['Regular_start'], 2)
    agg['Kakao_codes_pct'] = safe_pct(agg['Kakao_codes'], agg['Kakao_start'], 2)
    agg['Активность_24ч_pct'] = safe_pct(agg['Активен_24ч'], agg['Устройств'])
    agg['Нулевых_pct'] = safe_pct(agg['Нулевой_доход'], agg['Устройств'])
    agg['Мертвых_pct'] = safe_pct(agg['Мертвый'], agg['Устройств'])

    return agg


def compute_api_breakdown(all_data: pd.DataFrame, emp_totals: dict) -> pd.DataFrame:
    """Разбивка по API Key внутри сотрудника."""
    has_ext = 'Ext ID' in all_data.columns
    if has_ext:
        api = all_data.groupby(['_Сотрудник', 'API Key']).agg(
            Устройств=('Ext ID', 'nunique'),
            Доход=('Итого_руб', 'sum')
        ).reset_index()
    else:
        api = all_data.groupby(['_Сотрудник', 'API Key']).agg(
            Устройств=('_Сотрудник', 'count'),
            Доход=('Итого_руб', 'sum')
        ).reset_index()

    api['На_устр'] = safe_div(api['Доход'], api['Устройств'], 2)
    api['Доля_сотрудника'] = api.apply(
        lambda r: r['Доход'] / emp_totals.get(r['_Сотрудник'], 1) if emp_totals.get(r['_Сотрудник'], 0) else 0,
        axis=1
    )
    api = api.sort_values(['_Сотрудник', 'Доход'], ascending=[True, False]).reset_index(drop=True)
    return api


def compute_top_devices(all_data: pd.DataFrame, n: int = 30) -> pd.DataFrame:
    cols_keep = ['_Сотрудник', 'API Key']
    if 'Ext ID' in all_data.columns: cols_keep.append('Ext ID')
    if 'Operator' in all_data.columns: cols_keep.append('Operator')
    cols_keep += ['Итого_руб', 'Доход_Wa', 'Доход_Regular', 'Доход_Kakao']
    return all_data.nlargest(n, 'Итого_руб')[cols_keep].reset_index(drop=True)


def compute_model_stats(all_data: pd.DataFrame) -> pd.DataFrame:
    """Аналитика по моделям устройств."""
    if 'Operator' not in all_data.columns:
        return pd.DataFrame()
    all_data['_Model'] = all_data['Operator'].astype(str).str.split(' ver').str[0]
    device_agg = ('Ext ID', 'nunique') if 'Ext ID' in all_data.columns else ('_Model', 'count')
    g = all_data.groupby('_Model').agg(
        Устройств=device_agg,
        Доход=('Итого_руб', 'sum'),
        WA_sent=('Wa_sent', 'sum'),
        WA_delivered=('Wa_delivered', 'sum'),
        WA_accounts=('Wa_accounts', 'sum'),
        WA_bans=('Wa_bans', 'sum'),
    ).reset_index()
    g['На_устр'] = safe_div(g['Доход'], g['Устройств'], 2)
    g['WA_доставка_pct'] = safe_pct(g['WA_delivered'], g['WA_sent'])
    g['WA_ban_pct'] = safe_pct(g['WA_bans'], g['WA_accounts'])
    return g.sort_values('Доход', ascending=False).reset_index(drop=True)


# =============================================================
# АНАЛИЗ ЭФФЕКТИВНОСТИ ПО СЕРВИСАМ (динамический)
# =============================================================
# Современный отчёт из системы содержит десятки сервисов, у каждого
# единый набор из 12 метрик-счётчиков + 4 денежных:
#   <Сервис> Start / Number / Code / No Code / Waiting / Bad /
#            Error 1 / Error 2 / Account / Account Ban / Sent / Delivered
#   <сервис> Code Cost / Code Total / Sent Cost / Sent Total
#
# Функция автоматически находит ВСЕ сервисы по суффиксу 'Start Count'
# и считает метрики эффективности, о которых просил пользователь:
#   - среднее SMS с аккаунта     = Sent / Account
#   - % забаненных аккаунтов     = Account Ban / Account
#   - % регистраций от кодов     = Account / Code
#   - конверсия кода             = Code / Start
#   - % доставки                 = Delivered / Sent
#   - % брака                    = Bad / Start

# Метрики-счётчики, которые есть у каждого сервиса
SERVICE_COUNTERS = [
    'Start Count', 'Number Count', 'Code Count', 'No Code Count',
    'Waiting Count', 'Bad Count', 'Error 1 Count', 'Error 2 Count',
    'Account Count', 'Account Ban Count', 'Sent Count', 'Delivered Count',
]


def discover_services(all_data: pd.DataFrame) -> list:
    """
    Находит все сервисы в данных по наличию колонки '<Сервис> Start Count'.
    Возвращает список префиксов сервисов (без хвостового пробела).
    """
    services = []
    for col in all_data.columns:
        if col.endswith('Start Count'):
            prefix = col[:-len('Start Count')].rstrip()
            # Пропускаем пустой префикс (колонка ' Start Count' без имени сервиса)
            if prefix:
                services.append(prefix)
    return services


def _service_col(all_data, prefix, counter):
    """Безопасно достаёт колонку сервиса (учитывая возможный \\n и регистр)."""
    # Прямое совпадение
    candidates = [
        f'{prefix} {counter}',
        f'{prefix}\n {counter}',
        f'{prefix}\n{counter}',
    ]
    for c in candidates:
        if c in all_data.columns:
            return pd.to_numeric(all_data[c], errors='coerce').fillna(0)
    # Поиск по нормализованному имени
    target = normalize_key(f'{prefix}{counter}')
    for col in all_data.columns:
        if normalize_key(col) == target:
            return pd.to_numeric(all_data[col], errors='coerce').fillna(0)
    return pd.Series([0] * len(all_data), index=all_data.index)


def compute_service_efficiency(all_data: pd.DataFrame, by_employee: str = None) -> pd.DataFrame:
    """
    Считает метрики эффективности по каждому сервису.

    Если by_employee=None — агрегирует по всем данным (общая картина).
    Если by_employee='имя' — только по строкам этого сотрудника.
    """
    services = discover_services(all_data)
    data = all_data if by_employee is None else all_data[all_data['_Сотрудник'] == by_employee]

    rows = []
    for svc in services:
        start = _service_col(data, svc, 'Start Count').sum()
        number = _service_col(data, svc, 'Number Count').sum()
        code = _service_col(data, svc, 'Code Count').sum()
        bad = _service_col(data, svc, 'Bad Count').sum()
        account = _service_col(data, svc, 'Account Count').sum()
        ban = _service_col(data, svc, 'Account Ban Count').sum()
        sent = _service_col(data, svc, 'Sent Count').sum()
        delivered = _service_col(data, svc, 'Delivered Count').sum()

        # Пропускаем полностью пустые сервисы
        if (start + code + account + sent) == 0:
            continue

        def pct(a, b):
            return round(a / b * 100, 1) if b else 0.0

        # Флаг аномалии: счётчик брака/кодов превышает старты
        # (бывает когда в данных накопительные счётчики не синхронны)
        anomaly = ''
        if start > 0 and (bad > start or code > start):
            anomaly = 'счётчики рассинхронизированы (брак/коды > стартов)'

        rows.append({
            'Сервис': svc.replace('\n', '').strip(),
            'Старты': int(start),
            'Номера': int(number),
            'Коды': int(code),
            'Брак': int(bad),
            'Аккаунты': int(account),
            'Баны_акк': int(ban),
            'Отправлено': int(sent),
            'Доставлено': int(delivered),
            # === Метрики эффективности ===
            'Конверсия_кода_pct': pct(code, start),                 # код / старт
            'Брак_pct': pct(bad, start),                            # брак / старт
            'Регистраций_от_кодов_pct': pct(account, code),         # аккаунты / коды
            'Бан_аккаунтов_pct': pct(ban, account),                 # баны / аккаунты
            'Доставка_pct': pct(delivered, sent),                   # доставлено / отправлено
            'SMS_на_аккаунт': round(sent / account, 1) if account else 0.0,  # отправлено / аккаунт
            'Доставлено_на_аккаунт': round(delivered / account, 1) if account else 0.0,
            'Примечание': anomaly,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values('Старты', ascending=False).reset_index(drop=True)


def compute_service_by_employee(all_data: pd.DataFrame) -> pd.DataFrame:
    """
    Сводная таблица: сотрудник × сервис, ключевые метрики эффективности.
    Для понимания кто на каком сервисе насколько эффективен.
    """
    services = discover_services(all_data)
    employees = all_data['_Сотрудник'].unique()

    rows = []
    for emp in employees:
        data = all_data[all_data['_Сотрудник'] == emp]
        for svc in services:
            start = _service_col(data, svc, 'Start Count').sum()
            code = _service_col(data, svc, 'Code Count').sum()
            account = _service_col(data, svc, 'Account Count').sum()
            ban = _service_col(data, svc, 'Account Ban Count').sum()
            sent = _service_col(data, svc, 'Sent Count').sum()
            delivered = _service_col(data, svc, 'Delivered Count').sum()

            if (start + code + account + sent) == 0:
                continue

            def pct(a, b):
                return round(a / b * 100, 1) if b else 0.0

            rows.append({
                'Сотрудник': emp,
                'Сервис': svc.replace('\n', '').strip(),
                'Старты': int(start),
                'Коды': int(code),
                'Аккаунты': int(account),
                'Отправлено': int(sent),
                'Конверсия_кода_pct': pct(code, start),
                'Бан_аккаунтов_pct': pct(ban, account),
                'Доставка_pct': pct(delivered, sent),
                'SMS_на_аккаунт': round(sent / account, 1) if account else 0.0,
            })

    return pd.DataFrame(rows)


# =============================================================
# ПЕРСОНАЛЬНЫЕ ДАННЫЕ ПО СОТРУДНИКУ (для индивидуальных вкладок)
# =============================================================

def compute_employee_detail(all_data: pd.DataFrame, employee: str,
                            target_per_day: float = None,
                            period_days: float = None) -> dict:
    """
    Собирает всё для персональной вкладки сотрудника:
      - сводка (доход, устройства, выполнение цели)
      - разбивка по API Key (платформам)
      - разбивка по моделям устройств
      - список устройств (с пометками проблем)
      - эффективность по сервисам у этого сотрудника
      - автоматический разбор причин недотягивания
    """
    if target_per_day is None:
        target_per_day = TARGET_PER_DEVICE_DAY

    data = all_data[all_data['_Сотрудник'] == employee].copy()
    if data.empty:
        return None

    has_ext = 'Ext ID' in data.columns

    # Для UI-аналитики используем выбранный период. Timestamp — только fallback.
    if period_days is None and 'Timestamp' in data.columns:
        ts = pd.to_datetime(data['Timestamp'], errors='coerce')
        period_h = (ts.max() - ts.min()).total_seconds() / 3600
        period_days = max(period_h / 24, 1.0 / 24)
    elif period_days is None:
        period_days = 1.0

    total_revenue = data['Итого_руб'].sum()
    n_devices = data['Ext ID'].nunique() if has_ext else len(data)
    fact_per_day = total_revenue / n_devices / period_days if (n_devices and period_days) else 0
    plan = target_per_day * n_devices * period_days
    completion = total_revenue / plan * 100 if plan else 0

    # --- Разбивка по API Key (платформам) ---
    if has_ext:
        by_api = data.groupby('API Key').agg(
            Устройств=('Ext ID', 'nunique'),
            Доход=('Итого_руб', 'sum'),
        ).reset_index()
    else:
        by_api = data.groupby('API Key').agg(
            Устройств=('Итого_руб', 'count'),
            Доход=('Итого_руб', 'sum'),
        ).reset_index()
    by_api['Доход_на_устр'] = safe_div(by_api['Доход'], by_api['Устройств'], 1)
    by_api['Факт_сутки'] = safe_div(by_api['Доход'], by_api['Устройств'] * period_days, 1)
    by_api = by_api.sort_values('Доход', ascending=False).reset_index(drop=True)

    # --- Разбивка по моделям ---
    by_model = pd.DataFrame()
    if 'Operator' in data.columns and has_ext:
        data['_Model'] = data['Operator'].astype(str).str.split(' ver').str[0].str.strip()
        by_model = data.groupby('_Model').agg(
            Устройств=('Ext ID', 'nunique'),
            Доход=('Итого_руб', 'sum'),
        ).reset_index()
        by_model['Доход_на_устр'] = safe_div(by_model['Доход'], by_model['Устройств'], 1)
        by_model['Факт_сутки'] = safe_div(by_model['Доход'], by_model['Устройств'] * period_days, 1)
        by_model = by_model.sort_values('Факт_сутки', ascending=False).reset_index(drop=True)

    # --- Список устройств с пометками ---
    if has_ext:
        dev = data.groupby('Ext ID').agg(
            Доход=('Итого_руб', 'sum'),
            API=('API Key', 'first'),
            Модель=('Operator', 'first'),
            Простой_ч=('Timedelta_h', 'min'),
        ).reset_index()
    else:
        dev = data[['Ext ID', 'Итого_руб', 'API Key', 'Operator', 'Timedelta_h']].copy()
        dev.columns = ['Ext ID', 'Доход', 'API', 'Модель', 'Простой_ч']

    dev['Факт_сутки'] = (dev['Доход'] / period_days).round(1)
    dev['Модель'] = dev['Модель'].astype(str).str.split(' ver').str[0].str.strip()

    # Пометка проблемы для каждого устройства
    def device_flag(row):
        if row['Доход'] == 0 and pd.notna(row['Простой_ч']) and row['Простой_ч'] > 48:
            return 'мёртвое (нет дохода, простой >2 дн)'
        if row['Доход'] == 0:
            return 'нулевой доход'
        if row['Факт_сутки'] < target_per_day * 0.5:
            return 'ниже половины цели'
        if row['Факт_сутки'] < target_per_day:
            return 'ниже цели'
        return 'OK'

    dev['Статус'] = dev.apply(device_flag, axis=1)
    dev = dev.sort_values('Доход').reset_index(drop=True)  # худшие сверху

    # --- Эффективность по сервисам у этого сотрудника ---
    svc = compute_service_efficiency(all_data, by_employee=employee)

    # --- Автоматический разбор причин недотягивания ---
    reasons = analyze_underperformance(
        dev, by_api, by_model, svc, fact_per_day, target_per_day, completion
    )

    return {
        'employee': employee,
        'total_revenue': total_revenue,
        'n_devices': n_devices,
        'period_days': period_days,
        'fact_per_day': fact_per_day,
        'plan': plan,
        'completion': completion,
        'by_api': by_api,
        'by_model': by_model,
        'devices': dev,
        'services': svc,
        'reasons': reasons,
    }


def analyze_underperformance(dev, by_api, by_model, svc, fact_per_day,
                             target_per_day, completion) -> list:
    """
    Автоматически формулирует причины, почему сотрудник не дотягивает до плана.
    Возвращает список строк-наблюдений, отсортированных по важности.
    """
    reasons = []
    n_dev = len(dev)
    if n_dev == 0:
        return ['Нет устройств в данных']

    # Если план выполнен — отметим это
    if completion >= 100:
        reasons.append(f"✓ Цель выполнена на {completion:.0f}% — план достигнут, так держать!")
        # всё равно покажем точки роста ниже

    # 1. Мёртвые и нулевые устройства
    n_dead = (dev['Статус'] == 'мёртвое (нет дохода, простой >2 дн)').sum()
    n_zero = (dev['Доход'] == 0).sum()
    if n_dead > 0:
        lost = n_dead * target_per_day  # сколько ₽/сутки теряется
        reasons.append(
            f"🔴 {n_dead} 'мёртвых' устройств ({n_dead/n_dev*100:.0f}% парка) — "
            f"не работают и простаивают >2 дней. Это потеря ~{lost:.0f} ₽/сутки. "
            f"Проверить: связь, аккаунты, переустановка."
        )
    elif n_zero > 0:
        reasons.append(
            f"🟡 {n_zero} устройств ({n_zero/n_dev*100:.0f}%) с нулевым доходом — "
            f"включены, но ничего не приносят. Проверить сценарии и аккаунты."
        )

    # 2. Доля устройств ниже половины цели
    n_half = (dev['Факт_сутки'] < target_per_day * 0.5).sum()
    if n_half > n_dev * 0.5:
        reasons.append(
            f"🔴 Больше половины парка ({n_half} из {n_dev}) работает ниже половины цели "
            f"({target_per_day*0.5:.0f} ₽/устр/сутки). Проблема системная, не в отдельных устройствах."
        )

    # 3. Слабая платформа (API Key)
    if len(by_api) > 1:
        worst_api = by_api.nsmallest(1, 'Факт_сутки').iloc[0]
        best_api = by_api.nlargest(1, 'Факт_сутки').iloc[0]
        if worst_api['Факт_сутки'] < target_per_day * 0.6 and worst_api['Устройств'] >= 3:
            reasons.append(
                f"🟡 Платформа '{worst_api['API Key']}' тянет вниз: {worst_api['Факт_сутки']:.0f} ₽/устр/сутки "
                f"({int(worst_api['Устройств'])} устр). Для сравнения лучшая '{best_api['API Key']}' даёт "
                f"{best_api['Факт_сутки']:.0f} ₽/устр/сутки. Разобраться что не так с '{worst_api['API Key']}'."
            )

    # 4. Слабые модели
    if len(by_model) > 1:
        weak = by_model[by_model['Устройств'] >= 3].nsmallest(2, 'Факт_сутки')
        if not weak.empty and weak.iloc[0]['Факт_сутки'] < target_per_day:
            models_str = ', '.join(f"{r['_Model']} ({r['Факт_сутки']:.0f} ₽/устр/сутки)"
                                   for _, r in weak.iterrows())
            reasons.append(
                f"🟡 Слабые модели устройств: {models_str}. "
                f"Возможно, старое/слабое железо — рассмотреть замену или перераспределение."
            )

    # 5. Проблемы с сервисами (высокий бан / низкая доставка)
    if svc is not None and not svc.empty:
        for _, s in svc.iterrows():
            if s['Аккаунты'] >= 50 and s['Бан_аккаунтов_pct'] >= 20:
                reasons.append(
                    f"🔴 Сервис '{s['Сервис']}': бан аккаунтов {s['Бан_аккаунтов_pct']:.0f}% "
                    f"({s['Баны_акк']} из {s['Аккаунты']}). Аккаунты горят — снизить интенсивность, прогрев."
                )
            if s['Отправлено'] >= 100 and 0 < s['Доставка_pct'] < 80:
                reasons.append(
                    f"🟡 Сервис '{s['Сервис']}': доставка только {s['Доставка_pct']:.0f}%. "
                    f"Часть рассылки не доходит — проверить качество номеров/аккаунтов."
                )

    if not reasons:
        reasons.append("Явных системных проблем не видно — точечно подтянуть устройства ниже цели.")

    return reasons


# =============================================================
# ПУНКТ 1: SANITY CHECKS — проверки перед сдачей отчёта
# =============================================================
# Чек-лист по мотивам skill `validate-data` из anthropic/knowledge-work-plugins:
# https://github.com/anthropics/knowledge-work-plugins/blob/main/data/skills/validate-data/SKILL.md

def run_sanity_checks(all_data: pd.DataFrame, summary: pd.DataFrame,
                      files_info: list, mode_info: dict = None,
                      period_days: float = None) -> list:
    """
    Возвращает список проверок [(severity, category, name, status, message), ...]

    severity:  HIGH / MED / LOW   — насколько важна
    status:    PASS / WARN / FAIL — результат проверки
    """
    checks = []

    def add(severity, category, name, status, message):
        checks.append({
            'severity': severity,
            'category': category,
            'name': name,
            'status': status,
            'message': message,
        })

    # --- DATA QUALITY ---

    # 0a. Режим загрузки
    if mode_info:
        if mode_info['mode'] == 'config':
            add('LOW', 'Качество данных', 'Режим загрузки',
                'PASS',
                f"Используется конфиг-маппинг API Key → сотрудник "
                f"({len(mode_info['config']['all_employees'])} сотрудников в конфиге)")
        else:
            add('LOW', 'Качество данных', 'Режим загрузки',
                'PASS', 'Используется маппинг по имени файла (конфиг не найден)')

    # 0b. Несматченные API Key
    if mode_info and mode_info['mode'] == 'config':
        unmatched = mode_info.get('unmatched_keys', [])
        if unmatched:
            keys_str = ', '.join(unmatched[:10])
            if len(unmatched) > 10:
                keys_str += f' (и ещё {len(unmatched)-10})'
            add('HIGH', 'Качество данных', 'Несматченные API Key',
                'FAIL' if len(unmatched) > 5 else 'WARN',
                f'{len(unmatched)} API Key из данных не найдены в конфиге: {keys_str}. '
                f'Они попали в группу "Неизвестно" — поправь конфиг.')
        else:
            add('LOW', 'Качество данных', 'Несматченные API Key',
                'PASS', 'Все API Key из данных нашли своего сотрудника в конфиге')

        # 0c. Сотрудники без данных
        emp_with_data = set(summary['Сотрудник'])
        emp_in_config = set(mode_info['config']['all_employees'])
        no_data = emp_in_config - emp_with_data
        if no_data:
            add('MED', 'Качество данных', 'Сотрудники без данных',
                'WARN',
                f'{len(no_data)} сотрудников из конфига нет в данных: {", ".join(sorted(no_data))}. '
                f'Либо они не работают, либо их API Key переименованы.')

    # 1. Source verification — сколько файлов загружено
    n_files = len(files_info)
    if n_files == 0:
        add('HIGH', 'Качество данных', 'Источники',
            'FAIL', 'Не загружено ни одного файла')
    else:
        add('LOW', 'Качество данных', 'Источники',
            'PASS', f'Загружено {n_files} файл(ов)')

    # 2. Freshness — насколько свежие данные
    if 'Timestamp' in all_data.columns:
        ts = pd.to_datetime(all_data['Timestamp'], errors='coerce').dropna()
        if len(ts):
            age_days = (pd.Timestamp.now() - ts.max()).days
            if age_days > 7:
                add('MED', 'Качество данных', 'Свежесть данных',
                    'WARN', f'Самая свежая запись {age_days} дн. назад — данные могут быть устаревшими')
            else:
                add('LOW', 'Качество данных', 'Свежесть данных',
                    'PASS', f'Самая свежая запись {age_days} дн. назад')
        else:
            add('MED', 'Качество данных', 'Свежесть данных',
                'WARN', 'Нет валидного Timestamp в данных')
    else:
        add('LOW', 'Качество данных', 'Свежесть данных',
            'WARN', 'Колонка Timestamp отсутствует')

    # 3. Completeness — пропуски в ключевых колонках
    key_cols = ['API Key', 'Ext ID', 'Total Code Cost', 'Total Sent Cost']
    for col in key_cols:
        if col in all_data.columns:
            null_pct = all_data[col].isna().sum() / len(all_data) * 100
            if null_pct > 5:
                add('HIGH', 'Качество данных', f'Пропуски в "{col}"',
                    'FAIL', f'{null_pct:.1f}% строк без значения')
            elif null_pct > 0:
                add('LOW', 'Качество данных', f'Пропуски в "{col}"',
                    'PASS', f'{null_pct:.1f}% пропусков — в норме')

    # 4. Deduplication — нет ли дублей строк
    if 'Ext ID' in all_data.columns and 'API Key' in all_data.columns:
        dup_count = all_data.duplicated(subset=['API Key', 'Ext ID']).sum()
        if dup_count > 0:
            add('MED', 'Качество данных', 'Дубликаты строк',
                'WARN', f'{dup_count} строк-дубликатов (одна пара API Key + Ext ID встречается несколько раз)')
        else:
            add('LOW', 'Качество данных', 'Дубликаты строк',
                'PASS', 'Дубликатов не найдено')

    # 5. Filter verification — итоговые строки исключены
    # (если бы они остались, это бы дало двойной счёт — мы уже видели такое в Болонском)
    rows_with_api_key = all_data['API Key'].notna().sum()
    if rows_with_api_key == len(all_data):
        add('LOW', 'Качество данных', 'Итоговые строки',
            'PASS', 'Все строки без API Key исключены (нет двойного счёта)')

    # --- CALCULATION CHECKS ---

    # 6. Subtotals sum — складываются ли каналы в Total
    total_per_channel = (
        summary['Доход_Click'].sum() + summary['Доход_Wa'].sum() +
        summary['Доход_Business'].sum() + summary['Доход_Regular'].sum() +
        summary['Доход_Viber'].sum() + summary['Доход_Kakao'].sum()
    )
    total_overall = summary['Итого_руб'].sum()
    diff_pct = abs(total_per_channel - total_overall) / max(total_overall, 1) * 100
    if diff_pct < 1:
        add('HIGH', 'Расчёты', 'Сумма каналов = Total',
            'PASS', f'Сумма каналов {total_per_channel:,.2f} = Total {total_overall:,.2f} (расхождение {diff_pct:.2f}%)')
    elif diff_pct < 5:
        add('MED', 'Расчёты', 'Сумма каналов = Total',
            'WARN', f'Расхождение {diff_pct:.2f}% между суммой каналов и Total — возможно, не все каналы учтены')
    else:
        add('HIGH', 'Расчёты', 'Сумма каналов = Total',
            'FAIL', f'Расхождение {diff_pct:.2f}%! Сумма по каналам {total_per_channel:,.2f}, Total {total_overall:,.2f}')

    # 7. Magnitude — нет ли отрицательных значений
    neg_count = (all_data['Итого_руб'] < 0).sum()
    if neg_count > 0:
        add('HIGH', 'Расчёты', 'Отрицательные доходы',
            'FAIL', f'{neg_count} строк с отрицательным Итого — ошибка в данных')
    else:
        add('LOW', 'Расчёты', 'Отрицательные доходы',
            'PASS', 'Все доходы ≥ 0')

    # 8. Percentages в разумном диапазоне (0-100%)
    pct_cols = ['Click_bad_pct', 'WA_ban_pct', 'WA_delivery_pct', 'Reg_delivery_pct']
    over_100 = []
    for col in pct_cols:
        if col in summary.columns:
            over = (summary[col] > 100).sum()
            if over > 0:
                over_100.append(f'{col} ({over} сотр.)')
    if over_100:
        add('HIGH', 'Расчёты', 'Проценты в [0;100]',
            'FAIL', f'Проценты > 100 в: {", ".join(over_100)} — ошибка деления')
    else:
        add('LOW', 'Расчёты', 'Проценты в [0;100]',
            'PASS', 'Все процентные метрики в норме')

    # 9. Денoминатор корректен — нет ли деления на ноль с подменой
    # (например, у сотрудника 0 WA-аккаунтов, но в Wa_bans >0 — невозможно)
    impossible = ((summary['Wa_accounts'] == 0) & (summary['Wa_bans'] > 0)).sum()
    if impossible > 0:
        add('HIGH', 'Расчёты', 'Знаменатели',
            'FAIL', f'{impossible} сотрудников имеют WA-баны без WA-аккаунтов — нарушена логика')
    else:
        add('LOW', 'Расчёты', 'Знаменатели',
            'PASS', 'Невозможных соотношений не обнаружено')

    # --- REASONABLENESS CHECKS ---

    # 10. Order of magnitude — выручка в разумном диапазоне
    total = summary['Итого_руб'].sum()
    if total <= 0:
        add('HIGH', 'Разумность', 'Порядок величины',
            'FAIL', f'Общий доход {total} — нулевой или отрицательный')
    elif total < 100:
        add('MED', 'Разумность', 'Порядок величины',
            'WARN', f'Общий доход всего {total:.2f} ₽ — возможно, фильтры исключили нужные данные')
    else:
        add('LOW', 'Разумность', 'Порядок величины',
            'PASS', f'Общий доход {total:,.2f} ₽ — в разумном диапазоне')

    # 11. Edge cases — устройства с 0 кодов
    if 'Ext ID' in all_data.columns:
        zero_devices = (all_data['Итого_руб'] == 0).sum()
        total_devices = all_data['Ext ID'].nunique()
        zero_pct = zero_devices / total_devices * 100 if total_devices else 0
        if zero_pct > 30:
            add('MED', 'Разумность', 'Устройства с 0 ₽',
                'WARN', f'{zero_devices} устр. ({zero_pct:.1f}%) ничего не заработали — много "нулевых"')
        else:
            add('LOW', 'Разумность', 'Устройства с 0 ₽',
                'PASS', f'{zero_devices} устр. ({zero_pct:.1f}%) с 0 ₽ — в норме')

    # 12. Red flag: ровно круглые суммы (может указывать на дефолтное значение)
    suspiciously_round = (
        (summary['Итого_руб'] > 0) &
        (summary['Итого_руб'] == summary['Итого_руб'].round(-2))  # кратно 100
    ).sum()
    if suspiciously_round > 0:
        add('LOW', 'Разумность', 'Подозрительно круглые суммы',
            'WARN', f'{suspiciously_round} сотр. имеют доход, кратный 100 ₽ — проверить, не дефолтные ли значения')

    # 13. Период данных — полный или обрезанный
    if 'Timestamp' in all_data.columns:
        ts = pd.to_datetime(all_data['Timestamp'], errors='coerce').dropna()
        if len(ts):
            span_days = (ts.max() - ts.min()).days + 1
            if span_days < 2:
                add('MED', 'Разумность', 'Длина периода',
                    'WARN', f'Период данных {span_days} дн — слишком короткий для выводов о трендах')
            elif span_days < 7:
                add('MED', 'Разумность', 'Длина периода',
                    'WARN', f'Период {span_days} дн — выводы по отдельным сотрудникам имеют большой шум')
            else:
                add('LOW', 'Разумность', 'Длина периода',
                    'PASS', f'Период {span_days} дн — достаточно для сравнений')

    # --- ANALYTICAL PITFALLS ---

    # 14. Incomplete period comparison — все ли сотрудники наблюдались одинаково
    if period_days is not None:
        add('LOW', 'Аналитические ловушки', 'Сопоставимость периодов',
            'PASS',
            f'Для расчёта цели используется выбранный период {period_days:g} дн. '
            f'Timestamp используется только как последняя активность.')
    elif 'Timestamp' in all_data.columns:
        by_emp = all_data.groupby('_Сотрудник')['Timestamp'].agg(['min', 'max'])
        by_emp['span_h'] = (pd.to_datetime(by_emp['max']) - pd.to_datetime(by_emp['min'])).dt.total_seconds() / 3600
        max_span = by_emp['span_h'].max()
        min_span = by_emp['span_h'].min()
        if max_span > 0 and (max_span - min_span) / max_span > 0.5:
            add('HIGH', 'Аналитические ловушки', 'Сопоставимость периодов',
                'WARN',
                f'У сотрудников разные периоды наблюдения ({min_span:.0f}–{max_span:.0f} ч). '
                f'Сравнение по абсолютному доходу некорректно — используйте "Доход/устр" или нормируйте на длину периода.')
        else:
            add('LOW', 'Аналитические ловушки', 'Сопоставимость периодов',
                'PASS', f'Периоды наблюдения у всех сотрудников сопоставимы')

    # 15. Survivorship bias
    add('MED', 'Аналитические ловушки', 'Survivorship bias',
        'WARN',
        'В выгрузке только активно работающие устройства. Если у сотрудника железо физически отключено '
        '(а не показывает 0 ₽), оно вообще не попадёт в анализ — реальный парк может быть больше.')

    # 16. Average of averages
    naive_avg = summary['Доход_на_устр'].mean()
    correct_avg = summary['Итого_руб'].sum() / summary['Устройств'].sum() if summary['Устройств'].sum() else 0
    diff = abs(naive_avg - correct_avg)
    add('LOW', 'Аналитические ловушки', 'Average of averages',
        'PASS',
        f'Средний доход/устр посчитан корректно: total/total_dev = {correct_avg:.0f} ₽ '
        f'(не путать с mean(per_emp) = {naive_avg:.0f} ₽).')

    # 17. Sample size — маленькие парки дают шум
    small_emp = summary[summary['Устройств'] < 10]
    if len(small_emp) > 0:
        names = ', '.join(small_emp['Сотрудник'].head(5).tolist())
        add('MED', 'Аналитические ловушки', 'Малая выборка',
            'WARN',
            f'У {len(small_emp)} сотрудников меньше 10 устройств ({names}) — '
            f'процентные метрики (бан, доставка) ненадёжны на таких выборках.')

    # 18. Достижение цели по компании
    total_plan = summary['План_за_период'].sum()
    total_fact = summary['Итого_руб'].sum()
    if total_plan > 0:
        overall_pct = total_fact / total_plan * 100
        n_above = int((summary['Выполнение_pct'] >= 100).sum())
        n_below_50 = int((summary['Выполнение_pct'] < 50).sum())

        if overall_pct >= 100:
            add('LOW', 'Цели', 'Выполнение цели (общее)',
                'PASS',
                f'Компания в целом выполняет цель {TARGET_PER_DEVICE_DAY:.0f} ₽/устр/сутки: '
                f'факт {total_fact:,.0f} ₽ vs план {total_plan:,.0f} ₽ ({overall_pct:.1f}%). '
                f'{n_above} из {len(summary)} сотрудников перевыполняют.')
        elif overall_pct >= 70:
            add('MED', 'Цели', 'Выполнение цели (общее)',
                'WARN',
                f'Компания недовыполняет цель: {overall_pct:.1f}% от плана '
                f'({total_fact:,.0f} из {total_plan:,.0f} ₽). '
                f'Только {n_above} из {len(summary)} сотрудников вышли на план.')
        else:
            add('HIGH', 'Цели', 'Выполнение цели (общее)',
                'WARN',
                f'Существенный недобор: {overall_pct:.1f}% от плана. '
                f'{n_below_50} сотрудник(ов) ниже 50% плана. '
                f'Цель {TARGET_PER_DEVICE_DAY:.0f} ₽/устр/сутки требует пересмотра либо тактики, либо самой цели.')

    return checks


def assess_confidence(checks: list, period_info: str, summary: pd.DataFrame) -> dict:
    """
    Финальная оценка отчёта по 3-уровневой шкале:
      - Ready to share
      - Share with caveats
      - Needs revision
    """
    fails = [c for c in checks if c['status'] == 'FAIL']
    high_warns = [c for c in checks if c['status'] == 'WARN' and c['severity'] == 'HIGH']
    med_warns = [c for c in checks if c['status'] == 'WARN' and c['severity'] == 'MED']

    if fails:
        level = 'Needs revision'
        level_code = 'RED'
        summary_text = f'Обнаружены критические ошибки ({len(fails)} шт). Исправить перед презентацией.'
    elif high_warns:
        level = 'Share with caveats'
        level_code = 'YELLOW'
        summary_text = f'Отчёт корректен, но требует {len(high_warns)} важных оговорок при презентации.'
    elif med_warns:
        level = 'Share with caveats'
        level_code = 'YELLOW'
        summary_text = f'Отчёт корректен. {len(med_warns)} оговорок средней важности.'
    else:
        level = 'Ready to share'
        level_code = 'GREEN'
        summary_text = 'Все проверки пройдены. Отчёт готов к презентации.'

    # Caveats для презентации
    caveats = []
    for c in fails + high_warns + med_warns:
        caveats.append(f"[{c['severity']}] {c['name']}: {c['message']}")

    return {
        'level': level,
        'level_code': level_code,
        'summary': summary_text,
        'caveats': caveats,
        'n_pass': sum(1 for c in checks if c['status'] == 'PASS'),
        'n_warn': sum(1 for c in checks if c['status'] == 'WARN'),
        'n_fail': len(fails),
    }


# =============================================================
# ПУНКТ 4: METHODOLOGY.MD — отдельный файл с описанием расчётов
# =============================================================

METHODOLOGY_TEMPLATE = """# Methodology — методология расчётов

Сгенерировано: {generated}
Период данных: {period}

---

## Что считаем и из каких полей

### Доход (Итого_руб)
**Формула:** `Total Code Cost + Total Sent Cost` по каждой строке отчёта.
**Источник:** колонки `Total Code Cost`, `Total Sent Cost` исходных xlsx.
**Единица:** рубли.

### Доход по каналам (₽)

| Канал | Формула |
|---|---|
| Click | `click Code Total` |
| WhatsApp рассылка | `wa_load Sent Total` |
| Business | `business Code Total` |
| Regular (SMS) | `regular Code Total + regular Sent Total` |
| Viber | `viber Code Total + viber Sent Total` |
| Kakao | `kakao Code Total + kakao Sent Total` |

Сумма каналов должна совпадать с `Итого_руб` — это проверяется автоматически (см. отчёт о валидации).

### Размер парка (Устройств)
**Формула:** `nunique(Ext ID)` для каждого сотрудника.
То есть количество **уникальных идентификаторов устройств**, а не строк.
Если устройство присутствует в нескольких строках — оно считается один раз.

### Доход на устройство
**Формула:** `sum(Итого_руб) / nunique(Ext ID)` для каждого сотрудника.
**Это правильный способ** считать средний доход. Неправильно было бы взять среднее от уже усреднённых значений (см. «Average of averages» в pitfalls).

### Целевой показатель: 50 ₽/устр/сутки

**Цель:** каждое устройство должно приносить **50 ₽ дохода в сутки**.

| Метрика | Формула | Что показывает |
|---|---|---|
| Период наблюдения | `max(Timestamp) − min(Timestamp)` по сотруднику | Длина окна данных в часах/сутках |
| Факт ₽/устр/сутки | `Итого_руб / Устройств / Период_суток` | Сколько устройство в среднем зарабатывает в сутки |
| Отклонение/сутки | `Факт_руб_устр_сутки − 50` | + значит выше цели, − ниже |
| План за период | `50 × Устройств × Период_суток` | Сколько должно было быть заработано |
| Отклонение за период | `Итого_руб − План_за_период` | Сколько недополучено или сверх плана |
| Выполнение плана | `Итого_руб / План_за_период × 100` | % достижения цели |

**Важно про разные периоды наблюдения.** У сотрудников периоды наблюдения могут сильно различаться (например, в нашем случае от 40 минут до 4 дней). Поэтому нельзя считать единый план «50 × устройств × фиксированный период» — план **индивидуален** для каждого сотрудника и считается по его фактическому периоду наблюдения. Это означает, что:

- Сравнение по абсолютному доходу некорректно (тот, у кого выгрузка покрывает 4 дня, всегда «победит» того, у кого выгрузка покрывает день).
- Сравнение по **% выполнения плана** и по **₽/устр/сутки** — корректно и нормировано.
- Если период наблюдения у сотрудника меньше 1 часа — нормировка всё равно работает (используем минимум 1 час, чтобы не делить на ноль).

**Цвета в отчёте:**
- 🟢 Зелёный — ≥ 100% плана (цель достигнута)
- 🟡 Жёлтый — 50–99% плана (недовыполнение)
- 🔴 Красный — < 50% плана (серьёзный недобор)

**Где менять значение цели:** константа `TARGET_PER_DEVICE_DAY` в начале файла `analyze_reports.py`. После изменения весь отчёт пересчитается под новую цель.

### Активность за 24ч
**Формула:** `(Timedelta_h < 24)` для каждой строки → суммируется по сотруднику и делится на парк.
**Допущение:** `Timedelta` в исходных данных интерпретируется как **часы** простоя устройства с момента последней активности.

> ⚠️ ВАЖНО: Если в вашей системе Timedelta измеряется в секундах или минутах, пороги нужно пересчитать. См. блок «Чувствительность к допущениям» ниже.

### "Мёртвые" устройства
**Формула:** `(Итого_руб == 0) И (Timedelta_h > 48)`
То есть устройство одновременно:
- ничего не заработало за период наблюдения,
- и не подавало признаков жизни более 48 часов.

### Качественные метрики

| Метрика | Формула |
|---|---|
| Click брак, % | `Click Bad Count / Click Start Count × 100` |
| WA бан, % | `Wa Load Account Ban Count / Wa Load Account Count × 100` |
| WA доставка, % | `Wa Load Delivered Count / Wa Load Sent Count × 100` |
| Regular доставка, % | `Regular Delivered Count / Regular Sent Count × 100` |
| Kakao конверсия, % | `Kakao Code Count / Kakao Start Count × 100` |

При нулевом знаменателе результат принудительно равен 0 (не NaN, не ошибка).

---

## Обработка особых случаев

### Итоговые строки в файлах
Многие исходные файлы содержат финальную строку с пустым `API Key`, где зафиксирована сумма по всем строкам выше. Если её не исключить — получим **двойной счёт**.
**Решение:** строки с `API Key IS NULL` исключаются на этапе загрузки.

### Сборные файлы
Если в файле несколько разных `API Key` И в названии есть подсказка ("неучт", "сборн", "свалк", "разное"), либо в имени файла ≥3 слов и ≥3 API Key — файл автоматически разбивается на под-сотрудников по `API Key`.
Пример: `Неучтенка_свалка_бабай_оптимус_Серега.xlsx` → 3 сотрудника: `BABAI`, `SVALKA`, `OPTIMUS`.

### Отсутствующие колонки
Если в исходных данных нет какой-то колонки (например, нет Viber-данных), она трактуется как 0 — без падения скрипта.

---

## Чувствительность к допущениям

Эти параметры заложены константами в скрипте. При изменении бизнес-логики их надо подкручивать:

| Параметр | Текущее значение | Где менять |
|---|---|---|
| **Целевой доход** | **50 ₽/устр/сутки** | константа `TARGET_PER_DEVICE_DAY` в начале файла |
| Граница «активен / неактивен» | 24 часа | функция `enrich()`, строка с `Активен_24ч` |
| Граница «мёртвый» | 48 часов простоя | функция `enrich()`, строка с `Мертвый` |
| Порог «высокий бан WA» | ≥ 15% | функция `build_recommendations()` |
| Порог «отличная доставка WA» | ≥ 92% | функция `build_recommendations()` |
| Порог «сильный Kakao» | ≥ 18% конверсии | функция `build_recommendations()` |
| Малая выборка для качества | < 10 устройств | функция `run_sanity_checks()` |

---

## Аналитические ловушки, которые мы избегаем

(по мотивам skill `validate-data` из официального репозитория Anthropic
https://github.com/anthropics/knowledge-work-plugins)

### Average of averages
Средний доход на устройство всегда считается как `сумма_доходов / сумма_устройств`, а не как среднее от уже посчитанных средних. Эти числа могут различаться в разы при разном размере парков.

### Incomplete period comparison
Сравниваем сотрудников только по нормированным метрикам (доход/устр, % бана). Абсолютные суммы зависят от длины периода наблюдения — для разных сотрудников периоды могут отличаться, и валидация это проверяет.

### Survivorship bias
Анализируем только устройства, попавшие в выгрузку. Физически отключённые устройства в выгрузке не присутствуют — реальный парк сотрудника может быть больше зафиксированного.

### Selection bias
Сотрудник попадает в «топ» по доходу не потому что лучше работает, а возможно потому что у него больше парк. Поэтому ключевой ranking-метрикой является **доход на устройство**, а не абсолютная сумма.

---

## Что не делается и почему

- **Не прогнозируем будущий доход.** За 4–7 дней наблюдения нельзя построить тренд.
- **Не присваиваем причины снижения.** Например, «у Маратика просел доход потому что …» — для этого нужен второй период наблюдения и контроль над переменными.
- **Не оцениваем сотрудников по абсолютному вкладу.** У них разные парки, разные периоды наблюдения, разные сценарии.

---

## Воспроизводимость

Чтобы повторить анализ:
1. Скачать xlsx-файлы за тот же период
2. Сложить в одну папку (имя файла = имя сотрудника)
3. Запустить `python analyze_reports.py <папка> <папка_для_отчёта>`
4. Сверить общий доход с разделом «Sanity checks → Сумма каналов = Total»

Если общий доход совпал — все остальные срезы рассчитаны корректно.
"""


def build_recommendations(summary: pd.DataFrame) -> list:
    """Автогенерация рекомендаций на базе фактов."""
    recs = []
    if len(summary) == 0:
        return recs

    median_per_dev = summary['Доход_на_устр'].median()
    median_click_bad = summary['Click_bad_pct'].median()
    median_wa_ban = summary['WA_ban_pct'].median()

    for _, row in summary.iterrows():
        plus = []
        minus = []

        # === ЦЕЛЬ: 50 ₽/устр/сутки ===
        fact = row['Факт_руб_устр_сутки']
        plan_pct = row['Выполнение_pct']
        deviation = row['Отклонение_сутки']
        if plan_pct >= 100:
            plus.append(
                f"Цель выполнена: {fact:.0f} ₽/устр/сутки vs план {TARGET_PER_DEVICE_DAY:.0f} ₽ "
                f"(+{deviation:.0f} ₽, {plan_pct:.0f}% плана)"
            )
        elif plan_pct >= 70:
            minus.append(
                f"Недовыполнение цели: {fact:.0f} ₽/устр/сутки ({plan_pct:.0f}% плана), "
                f"не хватает {abs(deviation):.0f} ₽/устр/сутки до цели"
            )
        else:
            minus.append(
                f"Серьёзное недовыполнение: {fact:.0f} ₽/устр/сутки ({plan_pct:.0f}% плана) — "
                f"парк работает в {plan_pct/100:.1f}× ниже цели {TARGET_PER_DEVICE_DAY:.0f} ₽/устр/сутки"
            )

        # СИЛЬНЫЕ
        if row['Доход_на_устр'] >= median_per_dev * 1.3:
            plus.append(f"Высокая эффективность: {row['Доход_на_устр']:.0f} ₽/устр (выше медианы)")
        if row['WA_delivery_pct'] >= 92 and row['Wa_sent'] > 100:
            plus.append(f"Отличная доставка WA: {row['WA_delivery_pct']:.1f}%")
        if row['Kakao_codes_pct'] >= 18 and row['Kakao_start'] > 100:
            plus.append(f"Сильный Kakao: {row['Kakao_codes_pct']:.1f}% код/старт")
        if row['Click_bad_pct'] <= 30 and row['Click_total'] > 50:
            plus.append(f"Чистый Click-трафик: брак {row['Click_bad_pct']:.1f}%")
        if row['WA_ban_pct'] <= 8 and row['Wa_accounts'] > 100:
            plus.append(f"Бережно с аккаунтами WA: бан {row['WA_ban_pct']:.1f}%")
        if row['Активность_24ч_pct'] >= 35:
            plus.append(f"Высокая активность парка: {row['Активность_24ч_pct']:.0f}% за 24ч")

        # СЛАБЫЕ
        if row['Доход_на_устр'] < median_per_dev * 0.7 and row['Устройств'] > 10:
            minus.append(f"Низкая эффективность: {row['Доход_на_устр']:.0f} ₽/устр (ниже медианы)")
        if row['Click_bad_pct'] >= max(60, median_click_bad + 10) and row['Click_total'] > 50:
            minus.append(f"Высокий брак Click: {row['Click_bad_pct']:.1f}% — пересобрать сценарий или отключить")
        if row['WA_ban_pct'] >= max(15, median_wa_ban + 4) and row['Wa_accounts'] > 100:
            minus.append(f"Высокий бан WA-аккаунтов: {row['WA_ban_pct']:.1f}% — снизить частоту/прогрев")
        if row['Мертвых_pct'] >= 5:
            minus.append(f"Много 'мёртвых' устройств: {int(row['Мертвый'])} шт ({row['Мертвых_pct']:.1f}%) — разобрать или починить")
        if row['Активность_24ч_pct'] < 20 and row['Устройств'] > 10:
            minus.append(f"Низкая активность парка: {row['Активность_24ч_pct']:.0f}% за 24ч")
        if row['WA_delivery_pct'] < 85 and row['Wa_sent'] > 100:
            minus.append(f"Низкая доставка WA: {row['WA_delivery_pct']:.1f}%")
        if row['Итого_руб'] < 100 and row['Устройств'] > 5:
            minus.append("Практически не работает — проверить, не сломан ли сценарий")

        if not plus:
            plus.append("Стабильная работа без особых выбросов")
        if not minus:
            minus.append("Серьёзных проблем не выявлено — продолжать в том же духе")

        # ПУНКТ 3: pitfall-warning — оговорки о надёжности оценки
        caveat_parts = []
        if row['Устройств'] < 10:
            caveat_parts.append(f"малая выборка ({int(row['Устройств'])} устр) — % метрики ненадёжны")
        if row['Wa_accounts'] < 50 and (row['WA_ban_pct'] > 0 or row['WA_delivery_pct'] > 0):
            caveat_parts.append(f"мало WA-аккаунтов ({int(row['Wa_accounts'])}) — оценка бана/доставки шумная")
        if row['Click_total'] < 50 and row['Click_bad_pct'] > 0:
            caveat_parts.append(f"мало Click-стартов ({int(row['Click_total'])}) — оценка брака шумная")

        caveat = ' • '.join(caveat_parts) if caveat_parts else ''

        recs.append({
            'Сотрудник': row['Сотрудник'],
            'Доход': row['Итого_руб'],
            'Плюсы': ' • '.join(plus),
            'Минусы': ' • '.join(minus),
            'Caveat': caveat,
        })

    return recs


# =============================================================
# 3. ГЕНЕРАЦИЯ EXCEL
# =============================================================

def fmt_rub(v):
    try:
        return f"{float(v):,.2f} ₽".replace(',', ' ')
    except Exception:
        return ''

def fmt_int(v):
    try:
        return f"{int(v):,}".replace(',', ' ')
    except Exception:
        return ''


def write_excel(summary, api_break, top_dev, model_stats, recs, period_info,
                checks, confidence, svc_eff, all_data, out_path,
                period_days: float = None):
    """Создаём многолистовой Excel-файл."""
    wb = Workbook()

    # Стили
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill('solid', start_color='1F4E78')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='2E75B6')
    arial = Font(name='Arial', size=11)
    arial_bold = Font(name='Arial', size=11, bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')
    thin = Side(border_style='thin', color='B4B4B4')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_header_row(ws, row, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = center; c.border = border
        ws.row_dimensions[row].height = 32

    def make_title(ws, cols, text):
        last_col = get_column_letter(cols)
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = text
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 28

    # ---- ЛИСТ: Общие выводы ----
    ws = wb.active
    ws.title = 'Общие выводы'
    make_title(ws, 2, 'Краткое резюме')

    total_devices_raw, total_revenue, total_device_days, avg_period_days, company_fact_day = company_goal_totals(summary)
    total_devices = int(total_devices_raw)
    avg_per_dev = total_revenue / total_devices if total_devices else 0
    total_active = int(summary['Активен_24ч'].sum())
    total_dead = int(summary['Мертвый'].sum())

    # Доли каналов
    ch_wa = summary['Доход_Wa'].sum()
    ch_reg = summary['Доход_Regular'].sum()
    ch_kakao = summary['Доход_Kakao'].sum()
    ch_click = summary['Доход_Click'].sum()
    ch_biz = summary['Доход_Business'].sum()
    ch_viber = summary['Доход_Viber'].sum()
    def pct(x): return f"{x/total_revenue*100:.1f}%" if total_revenue else "0%"

    # Расчёт компанейских показателей цели
    total_plan = summary['План_за_период'].sum()
    company_pct = total_revenue / total_plan * 100 if total_plan else 0
    n_above_plan = int((summary['Выполнение_pct'] >= 100).sum())
    n_below_half = int((summary['Выполнение_pct'] < 50).sum())

    rows = [
        ('▌ СТАТУС ОТЧЁТА', ''),
        ('Уровень готовности', f"{confidence['level']}"),
        ('Резюме', confidence['summary']),
        ('Проверок пройдено', f"{confidence['n_pass']} PASS · {confidence['n_warn']} WARN · {confidence['n_fail']} FAIL"),
        ('', ''),
        ('Период данных', period_info),
        ('Общий доход', fmt_rub(total_revenue)),
        ('Сотрудников в отчёте', f"{len(summary)}"),
        ('Устройств всего', fmt_int(total_devices)),
        ('Средний доход на устройство', fmt_rub(avg_per_dev)),
        ('Активных за последние 24ч', f"{total_active} ({total_active/total_devices*100:.0f}% парка)" if total_devices else "0"),
        ('"Мёртвых" устройств (нет дохода и простой >2 дн)', f"{total_dead} ({total_dead/total_devices*100:.0f}% парка)" if total_devices else "0"),
        ('', ''),
        ('▌ ЦЕЛЬ: 50 ₽/устр/сутки', ''),
        ('Целевой показатель', f"{TARGET_PER_DEVICE_DAY:.0f} ₽ на устройство в сутки"),
        ('Факт по компании', f"{company_fact_day:.1f} ₽/устр/сутки"),
        ('Отклонение от цели', f"{company_fact_day - TARGET_PER_DEVICE_DAY:+.1f} ₽/устр/сутки"),
        ('План за период', fmt_rub(total_plan)),
        ('Факт за период', fmt_rub(total_revenue)),
        ('Выполнение плана', f"{company_pct:.1f}%"),
        ('Перевыполняют план', f"{n_above_plan} из {len(summary)} сотрудников"),
        ('Ниже 50% плана', f"{n_below_half} из {len(summary)} сотрудников"),
        ('', ''),
        ('▌ СТРУКТУРА ДОХОДА', ''),
        ('WhatsApp рассылки', f"{fmt_rub(ch_wa)} ({pct(ch_wa)})"),
        ('Regular (SMS-коды)', f"{fmt_rub(ch_reg)} ({pct(ch_reg)})"),
        ('Kakao', f"{fmt_rub(ch_kakao)} ({pct(ch_kakao)})"),
        ('Click', f"{fmt_rub(ch_click)} ({pct(ch_click)})"),
        ('Business', f"{fmt_rub(ch_biz)} ({pct(ch_biz)})"),
        ('Viber', f"{fmt_rub(ch_viber)} ({pct(ch_viber)})"),
        ('', ''),
        ('▌ ТОП-3 ПО ВЫПОЛНЕНИЮ ЦЕЛИ', ''),
    ]
    # Топ-3 по выполнению цели (а не по абсолютному доходу — это честнее)
    top_by_goal = summary.nlargest(3, 'Выполнение_pct')
    for _, r in top_by_goal.iterrows():
        rows.append((
            r['Сотрудник'],
            f"{r['Факт_руб_устр_сутки']:.0f} ₽/устр/сутки ({r['Выполнение_pct']:.0f}% плана, {int(r['Устройств'])} устр)"
        ))
    rows.append(('', ''))
    rows.append(('▌ АНТИ-ТОП ПО ВЫПОЛНЕНИЮ ЦЕЛИ', ''))
    bad_by_goal = summary[summary['Устройств'] >= 10].nsmallest(3, 'Выполнение_pct')
    for _, r in bad_by_goal.iterrows():
        rows.append((
            r['Сотрудник'],
            f"{r['Факт_руб_устр_сутки']:.0f} ₽/устр/сутки ({r['Выполнение_pct']:.0f}% плана, {int(r['Устройств'])} устр)"
        ))

    # Caveats (только важные — HIGH и MED)
    if confidence['caveats']:
        rows.append(('', ''))
        rows.append(('▌ ВАЖНЫЕ ОГОВОРКИ ДЛЯ ПРЕЗЕНТАЦИИ', ''))
        for cv in confidence['caveats'][:10]:  # топ-10 caveats
            rows.append(('•', cv))

    # Цвета бейджа Confidence
    badge_colors = {'GREEN': 'C6EFCE', 'YELLOW': 'FFEB9C', 'RED': 'FFC7CE'}
    badge_text_colors = {'GREEN': '006100', 'YELLOW': '9C5700', 'RED': '9C0006'}

    for idx, (k, v) in enumerate(rows):
        r = 3 + idx
        a = ws.cell(row=r, column=1, value=k)
        b = ws.cell(row=r, column=2, value=v)
        a.font = arial_bold if (k.startswith('▌') or k in ('Период данных', 'Общий доход')) else arial
        b.font = arial
        if k.startswith('▌'):
            a.fill = PatternFill('solid', start_color='1F4E78')
            a.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        elif k == 'Уровень готовности':
            color = badge_colors.get(confidence['level_code'], 'D9D9D9')
            text_color = badge_text_colors.get(confidence['level_code'], '000000')
            a.fill = PatternFill('solid', start_color=color)
            b.fill = PatternFill('solid', start_color=color)
            a.font = Font(name='Arial', size=11, bold=True, color=text_color)
            b.font = Font(name='Arial', size=11, bold=True, color=text_color)
        a.alignment = Alignment(vertical='top', wrap_text=True)
        b.alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[r].height = 22

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 65

    # ---- ЛИСТ: Валидация (sanity checks) ----
    ws = wb.create_sheet('Валидация')
    headers = ['Категория', 'Проверка', 'Статус', 'Важность', 'Подробности']
    make_title(ws, len(headers), f'Проверки качества отчёта · {confidence["level"]}')
    make_header_row(ws, 3, headers)

    # Цвета статусов
    status_fills = {
        'PASS': PatternFill('solid', start_color='C6EFCE'),
        'WARN': PatternFill('solid', start_color='FFEB9C'),
        'FAIL': PatternFill('solid', start_color='FFC7CE'),
    }
    status_fonts = {
        'PASS': Font(name='Arial', size=11, bold=True, color='006100'),
        'WARN': Font(name='Arial', size=11, bold=True, color='9C5700'),
        'FAIL': Font(name='Arial', size=11, bold=True, color='9C0006'),
    }
    severity_color = {'HIGH': 'C00000', 'MED': 'BF8F00', 'LOW': '595959'}

    # Сортировка: FAIL → HIGH WARN → MED WARN → LOW WARN → PASS
    status_order = {'FAIL': 0, 'WARN': 1, 'PASS': 2}
    severity_order = {'HIGH': 0, 'MED': 1, 'LOW': 2}
    sorted_checks = sorted(
        checks,
        key=lambda c: (status_order.get(c['status'], 9), severity_order.get(c['severity'], 9))
    )

    for idx, ch in enumerate(sorted_checks):
        r = 4 + idx
        ws.cell(row=r, column=1, value=ch['category']).alignment = left
        ws.cell(row=r, column=2, value=ch['name']).alignment = left
        ws.cell(row=r, column=2).font = arial_bold
        ws.cell(row=r, column=3, value=ch['status']).alignment = center
        ws.cell(row=r, column=3).fill = status_fills.get(ch['status'])
        ws.cell(row=r, column=3).font = status_fonts.get(ch['status'])
        ws.cell(row=r, column=4, value=ch['severity']).alignment = center
        ws.cell(row=r, column=4).font = Font(
            name='Arial', size=11, bold=True,
            color=severity_color.get(ch['severity'], '000000')
        )
        ws.cell(row=r, column=5, value=ch['message']).alignment = Alignment(
            wrap_text=True, vertical='center'
        )
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = border
            if ws.cell(row=r, column=col).font.bold is None:
                ws.cell(row=r, column=col).font = arial
        ws.row_dimensions[r].height = max(20, len(ch['message']) // 60 * 15 + 20)

    # Итоговая строка summary
    summary_row = 4 + len(sorted_checks) + 1
    ws.cell(row=summary_row, column=1, value='ИТОГО').font = arial_bold
    ws.cell(row=summary_row, column=2, value=confidence['level']).font = Font(
        name='Arial', size=12, bold=True,
        color=badge_text_colors.get(confidence['level_code'], '000000')
    )
    ws.cell(row=summary_row, column=2).fill = PatternFill(
        'solid', start_color=badge_colors.get(confidence['level_code'], 'D9D9D9')
    )
    ws.cell(row=summary_row, column=3, value=f"{confidence['n_pass']} PASS").font = status_fonts['PASS']
    ws.cell(row=summary_row, column=4, value=f"{confidence['n_warn']} WARN").font = status_fonts['WARN']
    ws.cell(row=summary_row, column=5, value=f"{confidence['n_fail']} FAIL").font = status_fonts['FAIL']
    for col in range(1, 6):
        ws.cell(row=summary_row, column=col).border = border

    for i, w in enumerate([22, 32, 10, 12, 70], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: Рейтинг ----
    ws = wb.create_sheet('Рейтинг')
    headers = ['Место','Сотрудник','Доход, ₽','Доля, %','Устройств','Доход/устр, ₽',
               'WhatsApp, ₽','Regular, ₽','Kakao, ₽','Click, ₽','Кодов']
    make_title(ws, len(headers), 'Рейтинг сотрудников по доходу')
    make_header_row(ws, 3, headers)

    total_row = 4 + len(summary)
    for idx, row in summary.iterrows():
        r = 4 + idx
        ws.cell(row=r, column=1, value=idx+1).alignment = center
        ws.cell(row=r, column=2, value=row['Сотрудник']).alignment = left
        ws.cell(row=r, column=3, value=row['Итого_руб']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=4, value=f'=C{r}/C{total_row}*100').number_format = '0.0"%"'
        ws.cell(row=r, column=5, value=int(row['Устройств'])).alignment = center
        ws.cell(row=r, column=6, value=f'=IF(E{r}=0,0,C{r}/E{r})').number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=7, value=row['Доход_Wa']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=8, value=row['Доход_Regular']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=9, value=row['Доход_Kakao']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=10, value=row['Доход_Click']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=11, value=int(row['Коды_всего'])).number_format = '#,##0'
        for col in range(1, 12):
            ws.cell(row=r, column=col).font = arial
            ws.cell(row=r, column=col).border = border

    # Итоговая строка
    ws.cell(row=total_row, column=1, value='ИТОГО').font = arial_bold
    ws.cell(row=total_row, column=3, value=f'=SUM(C4:C{total_row-1})').number_format = '#,##0.00 ₽'
    ws.cell(row=total_row, column=4, value=f'=SUM(D4:D{total_row-1})').number_format = '0.0"%"'
    ws.cell(row=total_row, column=5, value=f'=SUM(E4:E{total_row-1})')
    ws.cell(row=total_row, column=6, value=f'=IF(E{total_row}=0,0,C{total_row}/E{total_row})').number_format = '#,##0.00 ₽'
    ws.cell(row=total_row, column=7, value=f'=SUM(G4:G{total_row-1})').number_format = '#,##0.00 ₽'
    ws.cell(row=total_row, column=8, value=f'=SUM(H4:H{total_row-1})').number_format = '#,##0.00 ₽'
    ws.cell(row=total_row, column=9, value=f'=SUM(I4:I{total_row-1})').number_format = '#,##0.00 ₽'
    ws.cell(row=total_row, column=10, value=f'=SUM(J4:J{total_row-1})').number_format = '#,##0.00 ₽'
    ws.cell(row=total_row, column=11, value=f'=SUM(K4:K{total_row-1})').number_format = '#,##0'
    for col in range(1, 12):
        ws.cell(row=total_row, column=col).font = arial_bold
        ws.cell(row=total_row, column=col).fill = PatternFill('solid', start_color='DDEBF7')
        ws.cell(row=total_row, column=col).border = border

    ws.conditional_formatting.add(f'F4:F{total_row-1}', ColorScaleRule(
        start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'))
    ws.conditional_formatting.add(f'C4:C{total_row-1}', DataBarRule(
        start_type='min', end_type='max', color='4472C4'))

    for i, w in enumerate([8, 24, 16, 11, 12, 16, 16, 16, 16, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: Цели ----
    # Цель = TARGET_PER_DEVICE_DAY ₽ × устройств × дней наблюдения.
    # У сотрудников разные периоды наблюдения (валидатор это поднимает),
    # поэтому план строится индивидуально для каждого, а не как 50 ₽ × устр × фикс.период.
    ws = wb.create_sheet('Цели')
    headers = ['Сотрудник', 'Устройств', 'Период, сут',
               f'Цель, ₽/устр/сутки', 'Факт, ₽/устр/сутки', 'Отклонение, ₽/сутки',
               'План за период, ₽', 'Факт за период, ₽', 'Отклонение за период, ₽',
               'Выполнение, %']
    make_title(ws, len(headers), f'Достижение цели · {TARGET_PER_DEVICE_DAY:.0f} ₽ на устройство в сутки')
    make_header_row(ws, 3, headers)

    # Сортировка по проценту выполнения — лучшие сверху
    summary_by_goal = summary.sort_values('Выполнение_pct', ascending=False).reset_index(drop=True)

    for idx, row in summary_by_goal.iterrows():
        r = 4 + idx
        ws.cell(row=r, column=1, value=row['Сотрудник']).alignment = left
        ws.cell(row=r, column=2, value=int(row['Устройств'])).alignment = center
        ws.cell(row=r, column=3, value=row['Период_суток']).number_format = '0.00'
        ws.cell(row=r, column=3).alignment = center
        ws.cell(row=r, column=4, value=TARGET_PER_DEVICE_DAY).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=5, value=row['Факт_руб_устр_сутки']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=6, value=row['Отклонение_сутки']).number_format = '+#,##0.00 ₽;-#,##0.00 ₽;0.00 ₽'
        ws.cell(row=r, column=7, value=row['План_за_период']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=8, value=row['Итого_руб']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=9, value=row['Отклонение_за_период']).number_format = '+#,##0.00 ₽;-#,##0.00 ₽;0.00 ₽'
        ws.cell(row=r, column=10, value=row['Выполнение_pct']/100).number_format = '0.0%'

        # Подкрашиваем строку: зелёным >100%, красным <50%, жёлтым между
        pct = row['Выполнение_pct']
        if pct >= 100:
            row_fill = PatternFill('solid', start_color='E2EFDA')
        elif pct < 50:
            row_fill = PatternFill('solid', start_color='FCE4D6')
        else:
            row_fill = PatternFill('solid', start_color='FFF2CC')

        for col in range(1, 11):
            ws.cell(row=r, column=col).font = arial
            ws.cell(row=r, column=col).border = border
            if col in (4, 5, 6, 7, 8, 9, 10):
                ws.cell(row=r, column=col).alignment = right
            ws.cell(row=r, column=col).fill = row_fill

    # Итоговая строка
    total_row_goals = 4 + len(summary_by_goal)
    ws.cell(row=total_row_goals, column=1, value='ИТОГО по компании').font = arial_bold
    ws.cell(row=total_row_goals, column=2, value=int(summary['Устройств'].sum())).font = arial_bold
    total_devices, total_revenue, total_device_days, avg_period, company_fact_day = company_goal_totals(summary)
    ws.cell(row=total_row_goals, column=3, value=avg_period).number_format = '0.00'
    ws.cell(row=total_row_goals, column=3).font = arial_bold
    ws.cell(row=total_row_goals, column=4, value=TARGET_PER_DEVICE_DAY).number_format = '#,##0.00 ₽'
    ws.cell(row=total_row_goals, column=4).font = arial_bold
    # По компании средний факт = total / sum(устройств × период).
    total_plan_period = summary['План_за_период'].sum()
    ws.cell(row=total_row_goals, column=5, value=company_fact_day).number_format = '#,##0.00 ₽'
    ws.cell(row=total_row_goals, column=5).font = arial_bold
    ws.cell(row=total_row_goals, column=6, value=company_fact_day - TARGET_PER_DEVICE_DAY).number_format = '+#,##0.00 ₽;-#,##0.00 ₽;0.00 ₽'
    ws.cell(row=total_row_goals, column=6).font = arial_bold
    ws.cell(row=total_row_goals, column=7, value=total_plan_period).number_format = '#,##0.00 ₽'
    ws.cell(row=total_row_goals, column=7).font = arial_bold
    ws.cell(row=total_row_goals, column=8, value=total_revenue).number_format = '#,##0.00 ₽'
    ws.cell(row=total_row_goals, column=8).font = arial_bold
    ws.cell(row=total_row_goals, column=9, value=total_revenue - total_plan_period).number_format = '+#,##0.00 ₽;-#,##0.00 ₽;0.00 ₽'
    ws.cell(row=total_row_goals, column=9).font = arial_bold
    company_pct = total_revenue / total_plan_period if total_plan_period else 0
    ws.cell(row=total_row_goals, column=10, value=company_pct).number_format = '0.0%'
    ws.cell(row=total_row_goals, column=10).font = arial_bold
    for col in range(1, 11):
        ws.cell(row=total_row_goals, column=col).fill = PatternFill('solid', start_color='DDEBF7')
        ws.cell(row=total_row_goals, column=col).border = border

    # Цветовая шкала для выполнения %
    ws.conditional_formatting.add(f'J4:J{total_row_goals-1}', ColorScaleRule(
        start_type='num', start_value=0, start_color='F8696B',
        mid_type='num', mid_value=1.0, mid_color='FFEB84',
        end_type='num', end_value=1.5, end_color='63BE7B'))

    # Пояснение под таблицей
    note_row = total_row_goals + 2
    ws.cell(row=note_row, column=1,
            value='Как считается:').font = arial_bold
    ws.cell(row=note_row+1, column=1,
            value=f'• Цель — {TARGET_PER_DEVICE_DAY:.0f} ₽ на каждое устройство в сутки')
    ws.cell(row=note_row+2, column=1,
            value='• План за период = Цель × Устройств × Период в сутках')
    ws.cell(row=note_row+3, column=1,
            value='• В строке ИТОГО период — средний период, взвешенный по устройствам')
    ws.cell(row=note_row+4, column=1,
            value='• Зелёный = ≥100% плана; Жёлтый = 50-99%; Красный = <50%')
    for r in range(note_row, note_row+5):
        ws.cell(row=r, column=1).font = arial if r > note_row else arial_bold

    for i, w in enumerate([24, 12, 13, 17, 18, 20, 18, 18, 22, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: Эффективность сервисов ----
    if svc_eff is not None and not svc_eff.empty:
        ws = wb.create_sheet('Эффективность сервисов')
        headers = ['Сервис', 'Старты', 'Коды', 'Аккаунты', 'Баны акк.', 'Отправлено', 'Доставлено',
                   'Конверсия кода, %', 'Брак, %', 'Регистр. от кодов, %', 'Бан аккаунтов, %',
                   'Доставка, %', 'SMS на аккаунт']
        make_title(ws, len(headers), 'Эффективность по сервисам (вся компания)')
        make_header_row(ws, 3, headers)

        for idx, row in svc_eff.iterrows():
            r = 4 + idx
            ws.cell(row=r, column=1, value=row['Сервис']).alignment = left
            ws.cell(row=r, column=1).font = arial_bold
            ws.cell(row=r, column=2, value=row['Старты']).number_format = '#,##0'
            ws.cell(row=r, column=3, value=row['Коды']).number_format = '#,##0'
            ws.cell(row=r, column=4, value=row['Аккаунты']).number_format = '#,##0'
            ws.cell(row=r, column=5, value=row['Баны_акк']).number_format = '#,##0'
            ws.cell(row=r, column=6, value=row['Отправлено']).number_format = '#,##0'
            ws.cell(row=r, column=7, value=row['Доставлено']).number_format = '#,##0'
            ws.cell(row=r, column=8, value=row['Конверсия_кода_pct']/100).number_format = '0.0%'
            ws.cell(row=r, column=9, value=row['Брак_pct']/100).number_format = '0.0%'
            ws.cell(row=r, column=10, value=row['Регистраций_от_кодов_pct']/100).number_format = '0.0%'
            ws.cell(row=r, column=11, value=row['Бан_аккаунтов_pct']/100).number_format = '0.0%'
            ws.cell(row=r, column=12, value=row['Доставка_pct']/100).number_format = '0.0%'
            ws.cell(row=r, column=13, value=row['SMS_на_аккаунт']).number_format = '0.0'
            for col in range(1, 14):
                ws.cell(row=r, column=col).font = arial if col != 1 else arial_bold
                ws.cell(row=r, column=col).border = border
                if col >= 2:
                    ws.cell(row=r, column=col).alignment = right

        last_r = 3 + len(svc_eff)
        # Конверсия кода — зелёный хорошо
        ws.conditional_formatting.add(f'H4:H{last_r}', ColorScaleRule(
            start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'))
        # Брак — красный плохо
        ws.conditional_formatting.add(f'I4:I{last_r}', ColorScaleRule(
            start_type='min', start_color='63BE7B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='F8696B'))
        # Бан аккаунтов — красный плохо
        ws.conditional_formatting.add(f'K4:K{last_r}', ColorScaleRule(
            start_type='min', start_color='63BE7B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='F8696B'))
        # Доставка — зелёный хорошо
        ws.conditional_formatting.add(f'L4:L{last_r}', ColorScaleRule(
            start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'))

        # Пояснения
        note_r = last_r + 2
        notes = [
            ('Как читать метрики:', ''),
            ('Конверсия кода', 'Коды / Старты — какая доля попыток дала код'),
            ('Брак', 'Bad / Старты — доля бракованных попыток'),
            ('Регистр. от кодов', 'Аккаунты / Коды — сколько кодов превратилось в аккаунты'),
            ('Бан аккаунтов', 'Баны / Аккаунты — доля сгоревших аккаунтов'),
            ('Доставка', 'Доставлено / Отправлено — доходимость рассылки'),
            ('SMS на аккаунт', 'Отправлено / Аккаунты — сколько сообщений шлёт один аккаунт'),
        ]
        for i, (k, v) in enumerate(notes):
            ws.cell(row=note_r + i, column=1, value=k).font = arial_bold if i == 0 else arial
            ws.cell(row=note_r + i, column=2, value=v).font = arial

        for i, w in enumerate([18, 11, 10, 11, 11, 12, 12, 16, 10, 18, 16, 12, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: Качество ----
    ws = wb.create_sheet('Качество')
    headers = ['Сотрудник','Click брак, %','WA бан, %','WA доставка, %','Reg доставка, %',
               'Reg код/старт, %','Kakao код/старт, %','WA акк','WA банов','Reg кодов','Kakao кодов']
    make_title(ws, len(headers), 'Качество работы по сотрудникам')
    make_header_row(ws, 3, headers)

    for idx, row in summary.iterrows():
        r = 4 + idx
        ws.cell(row=r, column=1, value=row['Сотрудник']).alignment = left
        ws.cell(row=r, column=2, value=row['Click_bad_pct']/100).number_format = '0.0%'
        ws.cell(row=r, column=3, value=row['WA_ban_pct']/100).number_format = '0.0%'
        ws.cell(row=r, column=4, value=row['WA_delivery_pct']/100).number_format = '0.0%'
        ws.cell(row=r, column=5, value=row['Reg_delivery_pct']/100).number_format = '0.0%'
        ws.cell(row=r, column=6, value=row['Reg_codes_pct']/100).number_format = '0.00%'
        ws.cell(row=r, column=7, value=row['Kakao_codes_pct']/100).number_format = '0.00%'
        ws.cell(row=r, column=8, value=int(row['Wa_accounts'])).number_format = '#,##0'
        ws.cell(row=r, column=9, value=int(row['Wa_bans'])).number_format = '#,##0'
        ws.cell(row=r, column=10, value=int(row['Regular_codes'])).number_format = '#,##0'
        ws.cell(row=r, column=11, value=int(row['Kakao_codes'])).number_format = '#,##0'
        for col in range(1, 12):
            ws.cell(row=r, column=col).font = arial
            ws.cell(row=r, column=col).border = border

    last_r = 3 + len(summary)
    # Брак и бан — красный плохо (большое плохо)
    for col_letter in ['B','C']:
        ws.conditional_formatting.add(f'{col_letter}4:{col_letter}{last_r}', ColorScaleRule(
            start_type='min', start_color='63BE7B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='F8696B'))
    # Доставки и конверсии — зелёный хорошо
    for col_letter in ['D','E','F','G']:
        ws.conditional_formatting.add(f'{col_letter}4:{col_letter}{last_r}', ColorScaleRule(
            start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'))

    for i, w in enumerate([24, 14, 14, 15, 16, 17, 18, 12, 12, 12, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: Утилизация ----
    ws = wb.create_sheet('Утилизация')
    headers = ['Сотрудник','Устройств','Активных 24ч','% активных','Нулевых','% нулевых','Мёртвых','% мёртвых','Доход, ₽']
    make_title(ws, len(headers), 'Утилизация устройств')
    make_header_row(ws, 3, headers)

    for idx, row in summary.iterrows():
        r = 4 + idx
        ws.cell(row=r, column=1, value=row['Сотрудник']).alignment = left
        ws.cell(row=r, column=2, value=int(row['Устройств'])).alignment = center
        ws.cell(row=r, column=3, value=int(row['Активен_24ч'])).alignment = center
        ws.cell(row=r, column=4, value=row['Активность_24ч_pct']/100).number_format = '0.0%'
        ws.cell(row=r, column=5, value=int(row['Нулевой_доход'])).alignment = center
        ws.cell(row=r, column=6, value=row['Нулевых_pct']/100).number_format = '0.0%'
        ws.cell(row=r, column=7, value=int(row['Мертвый'])).alignment = center
        ws.cell(row=r, column=8, value=row['Мертвых_pct']/100).number_format = '0.0%'
        ws.cell(row=r, column=9, value=row['Итого_руб']).number_format = '#,##0.00 ₽'
        for col in range(1, 10):
            ws.cell(row=r, column=col).font = arial
            ws.cell(row=r, column=col).border = border

    last_r = 3 + len(summary)
    ws.conditional_formatting.add(f'D4:D{last_r}', ColorScaleRule(
        start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'))
    ws.conditional_formatting.add(f'H4:H{last_r}', ColorScaleRule(
        start_type='min', start_color='63BE7B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='F8696B'))

    for i, w in enumerate([24, 14, 14, 13, 12, 12, 12, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: По API Key ----
    ws = wb.create_sheet('По API Key')
    headers = ['Сотрудник','API Key','Устройств','Доход, ₽','Доход/устр, ₽','Доля сотрудника']
    make_title(ws, len(headers), 'Эффективность по API Key (платформам)')
    make_header_row(ws, 3, headers)

    for idx, row in api_break.iterrows():
        r = 4 + idx
        ws.cell(row=r, column=1, value=row['_Сотрудник']).alignment = left
        ws.cell(row=r, column=2, value=row['API Key']).alignment = left
        ws.cell(row=r, column=3, value=int(row['Устройств'])).alignment = center
        ws.cell(row=r, column=4, value=row['Доход']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=5, value=f'=IF(C{r}=0,0,D{r}/C{r})').number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=6, value=row['Доля_сотрудника']).number_format = '0.0%'
        for col in range(1, 7):
            ws.cell(row=r, column=col).font = arial
            ws.cell(row=r, column=col).border = border

    if len(api_break) > 0:
        last_r = 3 + len(api_break)
        ws.conditional_formatting.add(f'E4:E{last_r}', ColorScaleRule(
            start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'))

    for i, w in enumerate([24, 18, 12, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: Топ-устройства ----
    ws = wb.create_sheet('Топ-устройства')
    headers = ['#','Сотрудник','API Key','Ext ID','Модель','Доход, ₽','WA, ₽']
    make_title(ws, len(headers), 'Топ устройств по доходу')
    make_header_row(ws, 3, headers)

    for idx, row in top_dev.iterrows():
        r = 4 + idx
        ws.cell(row=r, column=1, value=idx+1).alignment = center
        ws.cell(row=r, column=2, value=row['_Сотрудник']).alignment = left
        ws.cell(row=r, column=3, value=row['API Key']).alignment = left
        ws.cell(row=r, column=4, value=str(row.get('Ext ID', ''))).alignment = left
        op = str(row.get('Operator', ''))
        ws.cell(row=r, column=5, value=op.split(' - ')[0] if op else '').alignment = left
        ws.cell(row=r, column=6, value=row['Итого_руб']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=7, value=row['Доход_Wa']).number_format = '#,##0.00 ₽'
        for col in range(1, 8):
            ws.cell(row=r, column=col).font = arial
            ws.cell(row=r, column=col).border = border

    for i, w in enumerate([6, 20, 16, 22, 28, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: Рекомендации ----
    ws = wb.create_sheet('Рекомендации')
    headers = ['Сотрудник','Доход, ₽','Сильные стороны','Что улучшить','Оговорки о надёжности']
    make_title(ws, len(headers), 'Персональные рекомендации')
    make_header_row(ws, 3, headers)

    for idx, rec in enumerate(recs):
        r = 4 + idx
        ws.cell(row=r, column=1, value=rec['Сотрудник']).font = arial_bold
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=r, column=2, value=rec['Доход']).number_format = '#,##0.00 ₽'
        ws.cell(row=r, column=2).alignment = Alignment(vertical='top')
        ws.cell(row=r, column=3, value=rec['Плюсы']).alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=r, column=3).fill = PatternFill('solid', start_color='E2EFDA')
        ws.cell(row=r, column=4, value=rec['Минусы']).alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=r, column=4).fill = PatternFill('solid', start_color='FCE4D6')
        caveat = rec.get('Caveat', '')
        ws.cell(row=r, column=5, value=caveat or '—').alignment = Alignment(vertical='top', wrap_text=True)
        if caveat:
            ws.cell(row=r, column=5).fill = PatternFill('solid', start_color='FFF2CC')
        for col in range(1, 6):
            ws.cell(row=r, column=col).font = arial if col != 1 else arial_bold
            ws.cell(row=r, column=col).border = border
        ws.row_dimensions[r].height = 80

    for i, w in enumerate([20, 14, 50, 55, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ЛИСТ: Модели устройств (если есть данные) ----
    if len(model_stats) > 0:
        ws = wb.create_sheet('Модели')
        headers = ['Модель','Устройств','Доход, ₽','Доход/устр, ₽','WA доставка, %','WA бан, %']
        make_title(ws, len(headers), 'Эффективность по моделям устройств')
        make_header_row(ws, 3, headers)

        for idx, row in model_stats.iterrows():
            r = 4 + idx
            ws.cell(row=r, column=1, value=row['_Model']).alignment = left
            ws.cell(row=r, column=2, value=int(row['Устройств'])).alignment = center
            ws.cell(row=r, column=3, value=row['Доход']).number_format = '#,##0.00 ₽'
            ws.cell(row=r, column=4, value=row['На_устр']).number_format = '#,##0.00 ₽'
            ws.cell(row=r, column=5, value=row['WA_доставка_pct']/100).number_format = '0.0%'
            ws.cell(row=r, column=6, value=row['WA_ban_pct']/100).number_format = '0.0%'
            for col in range(1, 7):
                ws.cell(row=r, column=col).font = arial
                ws.cell(row=r, column=col).border = border

        last_r = 3 + len(model_stats)
        ws.conditional_formatting.add(f'D4:D{last_r}', ColorScaleRule(
            start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'))

        for i, w in enumerate([20, 12, 16, 16, 16, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- ПЕРСОНАЛЬНЫЕ ЛИСТЫ ПО СОТРУДНИКАМ ----
    # По одному листу на сотрудника, чтобы каждый видел только своё
    # и понимал причины недотягивания до плана.
    def safe_sheet_name(name):
        # Excel: имя листа ≤31 символ, без : \ / ? * [ ]
        clean = re.sub(r'[:\\/?*\[\]]', '_', str(name))
        return ('👤 ' + clean)[:31]

    badge_colors_emp = {'GREEN': 'C6EFCE', 'YELLOW': 'FFEB9C', 'RED': 'FFC7CE'}
    badge_text_emp = {'GREEN': '006100', 'YELLOW': '9C5700', 'RED': '9C0006'}

    for emp_name in summary['Сотрудник']:
        detail = compute_employee_detail(all_data, emp_name, period_days=period_days)
        if detail is None:
            continue

        ws = wb.create_sheet(safe_sheet_name(emp_name))

        # Цвет по выполнению
        comp = detail['completion']
        code_color = 'GREEN' if comp >= 100 else ('YELLOW' if comp >= 50 else 'RED')

        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = f'{emp_name} — персональный разбор'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 28

        r = 3
        # Сводка
        ws.cell(row=r, column=1, value='Доход за период').font = arial_bold
        ws.cell(row=r, column=2, value=detail['total_revenue']).number_format = '#,##0.00 ₽'
        ws.cell(row=r+1, column=1, value='Устройств').font = arial_bold
        ws.cell(row=r+1, column=2, value=detail['n_devices'])
        ws.cell(row=r+2, column=1, value='Период наблюдения').font = arial_bold
        ws.cell(row=r+2, column=2, value=f"{detail['period_days']:.1f} суток")
        ws.cell(row=r+3, column=1, value='Факт ₽/устр/сутки').font = arial_bold
        ws.cell(row=r+3, column=2, value=detail['fact_per_day']).number_format = '#,##0.0 ₽'
        ws.cell(row=r+4, column=1, value=f'Цель ₽/устр/сутки').font = arial_bold
        ws.cell(row=r+4, column=2, value=TARGET_PER_DEVICE_DAY).number_format = '#,##0 ₽'
        ws.cell(row=r+5, column=1, value='Выполнение плана').font = arial_bold
        cc = ws.cell(row=r+5, column=2, value=comp/100)
        cc.number_format = '0.0%'
        cc.fill = PatternFill('solid', start_color=badge_colors_emp[code_color])
        cc.font = Font(name='Arial', size=11, bold=True, color=badge_text_emp[code_color])

        # Причины
        r = 10
        ws.cell(row=r, column=1, value='ПОЧЕМУ НЕ ДОТЯГИВАЕМ ДО ПЛАНА').font = Font(
            name='Arial', size=12, bold=True, color='FFFFFF')
        ws.cell(row=r, column=1).fill = PatternFill('solid', start_color='C00000')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        for reason in detail['reasons']:
            ws.cell(row=r, column=1, value=reason).alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = max(30, len(reason) // 80 * 15 + 20)
            r += 1

        # Платформы
        r += 1
        ws.cell(row=r, column=1, value='ПЛАТФОРМЫ (API Key)').font = arial_bold
        ws.cell(row=r, column=1).fill = PatternFill('solid', start_color='DDEBF7')
        r += 1
        api_headers = ['API Key', 'Устройств', 'Доход, ₽', '₽/устр', '₽/устр/сутки']
        for i, h in enumerate(api_headers, 1):
            c = ws.cell(row=r, column=i, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
        r += 1
        for _, ar in detail['by_api'].iterrows():
            ws.cell(row=r, column=1, value=ar['API Key']).border = border
            ws.cell(row=r, column=2, value=int(ar['Устройств'])).border = border
            ws.cell(row=r, column=3, value=ar['Доход']).number_format = '#,##0.00 ₽'
            ws.cell(row=r, column=3).border = border
            ws.cell(row=r, column=4, value=ar['Доход_на_устр']).number_format = '#,##0.0 ₽'
            ws.cell(row=r, column=4).border = border
            fs = ws.cell(row=r, column=5, value=ar['Факт_сутки'])
            fs.number_format = '#,##0.0 ₽'
            fs.border = border
            # подсветка платформы по факту/сутки
            if ar['Факт_сутки'] >= TARGET_PER_DEVICE_DAY:
                fill = 'E2EFDA'
            elif ar['Факт_сутки'] < TARGET_PER_DEVICE_DAY * 0.5:
                fill = 'FCE4D6'
            else:
                fill = 'FFF2CC'
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=fill)
            r += 1

        # Сервисы сотрудника
        if detail['services'] is not None and not detail['services'].empty:
            r += 1
            ws.cell(row=r, column=1, value='ЭФФЕКТИВНОСТЬ СЕРВИСОВ').font = arial_bold
            ws.cell(row=r, column=1).fill = PatternFill('solid', start_color='DDEBF7')
            r += 1
            svc_headers = ['Сервис', 'Старты', 'Коды', 'Бан акк %', 'Доставка %', 'SMS/акк']
            for i, h in enumerate(svc_headers, 1):
                c = ws.cell(row=r, column=i, value=h)
                c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
            r += 1
            for _, sr in detail['services'].iterrows():
                ws.cell(row=r, column=1, value=sr['Сервис']).border = border
                ws.cell(row=r, column=2, value=sr['Старты']).number_format = '#,##0'
                ws.cell(row=r, column=2).border = border
                ws.cell(row=r, column=3, value=sr['Коды']).number_format = '#,##0'
                ws.cell(row=r, column=3).border = border
                ws.cell(row=r, column=4, value=sr['Бан_аккаунтов_pct']/100).number_format = '0.0%'
                ws.cell(row=r, column=4).border = border
                ws.cell(row=r, column=5, value=sr['Доставка_pct']/100).number_format = '0.0%'
                ws.cell(row=r, column=5).border = border
                ws.cell(row=r, column=6, value=sr['SMS_на_аккаунт']).number_format = '0.0'
                ws.cell(row=r, column=6).border = border
                r += 1

        # Список устройств (худшие сверху)
        r += 1
        ws.cell(row=r, column=1, value='УСТРОЙСТВА (худшие сверху — с них начинать)').font = arial_bold
        ws.cell(row=r, column=1).fill = PatternFill('solid', start_color='DDEBF7')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        dev_headers = ['Ext ID', 'Платформа', 'Модель', 'Доход, ₽', '₽/сутки', 'Статус']
        for i, h in enumerate(dev_headers, 1):
            c = ws.cell(row=r, column=i, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
        r += 1
        # Показываем все устройства (отсортированы — худшие сверху)
        status_fill_map = {
            'мёртвое (нет дохода, простой >2 дн)': 'FFC7CE',
            'нулевой доход': 'FCE4D6',
            'ниже половины цели': 'FFF2CC',
            'ниже цели': 'FFF9E6',
            'OK': 'E2EFDA',
        }
        for _, dr in detail['devices'].iterrows():
            ws.cell(row=r, column=1, value=str(dr['Ext ID'])).border = border
            ws.cell(row=r, column=2, value=str(dr['API'])).border = border
            ws.cell(row=r, column=3, value=str(dr['Модель'])).border = border
            ws.cell(row=r, column=4, value=dr['Доход']).number_format = '#,##0.00 ₽'
            ws.cell(row=r, column=4).border = border
            ws.cell(row=r, column=5, value=dr['Факт_сутки']).number_format = '#,##0.0 ₽'
            ws.cell(row=r, column=5).border = border
            ws.cell(row=r, column=6, value=dr['Статус']).border = border
            fill = status_fill_map.get(dr['Статус'], 'FFFFFF')
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=fill)
            r += 1

        for i, w in enumerate([22, 14, 22, 14, 12, 36], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"  Excel: {out_path}")


# =============================================================
# 4. ГЕНЕРАЦИЯ HTML
# =============================================================

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>Отчёт по сотрудникам</title>
<style>
  * { box-sizing: border-box; }
  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
         margin: 0; padding: 24px; background: #f5f6f8; color: #1a1a1a; }
  .container { max-width: 1280px; margin: 0 auto; }
  h1 { margin: 0 0 8px; font-size: 28px; font-weight: 600; }
  .subtitle { color: #666; margin-bottom: 24px; }
  .cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 28px; }
  .card { background: #fff; border-radius: 10px; padding: 16px 20px; border: 1px solid #e5e7eb; }
  .card .label { font-size: 13px; color: #6b7280; }
  .card .value { font-size: 26px; font-weight: 600; margin-top: 4px; }
  .card .hint { font-size: 12px; color: #9ca3af; margin-top: 4px; }
  section { background: #fff; border-radius: 10px; border: 1px solid #e5e7eb; padding: 24px; margin-bottom: 24px; }
  section h2 { margin: 0 0 16px; font-size: 20px; font-weight: 600; }
  table { width: 100%; border-collapse: collapse; font-size: 14px; }
  th, td { padding: 10px 12px; text-align: left; border-bottom: 1px solid #f0f0f0; }
  th { background: #f9fafb; font-weight: 600; color: #374151; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px; }
  th.num, td.num { text-align: right; font-variant-numeric: tabular-nums; }
  tr:hover { background: #fafbfc; }
  .bar { background: linear-gradient(to right, #4472C4 var(--w), transparent var(--w)); }
  .good { color: #137333; font-weight: 500; }
  .bad { color: #c5221f; font-weight: 500; }
  .neutral { color: #5f6368; }
  .chip { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 12px; font-weight: 500; }
  .chip-green { background: #e6f4ea; color: #137333; }
  .chip-red { background: #fce8e6; color: #c5221f; }
  .chip-yellow { background: #fef7e0; color: #b06000; }
  .chip-blue { background: #e8f0fe; color: #1967d2; }
  .rec-grid { display: grid; grid-template-columns: 1fr; gap: 16px; }
  .rec { border: 1px solid #e5e7eb; border-radius: 8px; padding: 16px; background: #fafbfc; }
  .rec-head { display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 10px; }
  .rec-name { font-weight: 600; font-size: 16px; }
  .rec-money { color: #1967d2; font-weight: 600; }
  .rec-pluses { background: #e6f4ea; border-radius: 6px; padding: 8px 12px; margin-bottom: 6px; font-size: 13px; }
  .rec-minuses { background: #fce8e6; border-radius: 6px; padding: 8px 12px; font-size: 13px; }
  .channel-bar { display: flex; height: 24px; border-radius: 4px; overflow: hidden; margin-top: 8px; min-width: 200px; }
  .channel-bar > div { height: 100%; }
  .legend { display: flex; flex-wrap: wrap; gap: 16px; margin-bottom: 16px; font-size: 13px; color: #6b7280; }
  .legend > span { display: flex; align-items: center; gap: 6px; }
  .legend i { width: 12px; height: 12px; border-radius: 2px; display: inline-block; }
  .footer { text-align: center; color: #9ca3af; font-size: 12px; padding: 16px 0; }
  /* Employee cards */
  .emp-card { border: 1px solid #e5e7eb; border-radius: 8px; margin-bottom: 12px; overflow: hidden; }
  .emp-card summary { padding: 14px 18px; cursor: pointer; font-weight: 600; font-size: 15px; display: flex; justify-content: space-between; align-items: center; list-style: none; }
  .emp-card summary::-webkit-details-marker { display: none; }
  .emp-card summary:hover { background: #fafbfc; }
  .emp-card[open] summary { border-bottom: 1px solid #e5e7eb; }
  .emp-badge { font-size: 13px; font-weight: 700; padding: 4px 10px; border-radius: 6px; }
  .emp-badge.green { background: #e6f4ea; color: #137333; }
  .emp-badge.yellow { background: #fef7e0; color: #9c5700; }
  .emp-badge.red { background: #fce8e6; color: #c5221f; }
  .emp-body { padding: 16px 18px; }
  .emp-body h4 { margin: 16px 0 8px; font-size: 14px; color: #374151; }
  .emp-body h4:first-child { margin-top: 0; }
  .emp-reasons { list-style: none; padding: 0; margin: 0 0 8px; }
  .emp-reasons li { padding: 8px 12px; margin-bottom: 6px; border-radius: 6px; font-size: 13px; background: #f8f9fa; border-left: 3px solid #dadce0; }
  .emp-stats { display: flex; flex-wrap: wrap; gap: 16px; margin-bottom: 12px; font-size: 13px; }
  .emp-stats div { background: #f8f9fa; padding: 8px 12px; border-radius: 6px; }
  .emp-stats b { display: block; font-size: 18px; margin-top: 2px; }
  table.mini { font-size: 12px; }
  table.mini th, table.mini td { padding: 6px 8px; }
  tr.dev-dead { background: #fce8e6; }
  tr.dev-zero { background: #fdf0e8; }
  tr.dev-low { background: #fef9e8; }
  tr.dev-ok { background: #f1f8f4; }
  .tabs { display: flex; gap: 4px; margin-bottom: 16px; border-bottom: 1px solid #e5e7eb; }
  .tab { padding: 8px 16px; cursor: pointer; border-bottom: 2px solid transparent; font-size: 14px; color: #6b7280; }
  .tab.active { color: #1967d2; border-bottom-color: #1967d2; font-weight: 500; }
  /* Confidence badge */
  .confidence { padding: 16px 20px; border-radius: 10px; margin-bottom: 24px; display: flex; gap: 16px; align-items: center; }
  .confidence.green { background: #e6f4ea; border: 1px solid #34a853; }
  .confidence.yellow { background: #fef7e0; border: 1px solid #f9ab00; }
  .confidence.red { background: #fce8e6; border: 1px solid #ea4335; }
  .confidence .badge { font-size: 14px; font-weight: 700; padding: 6px 12px; border-radius: 6px; white-space: nowrap; }
  .confidence.green .badge { background: #34a853; color: #fff; }
  .confidence.yellow .badge { background: #f9ab00; color: #fff; }
  .confidence.red .badge { background: #ea4335; color: #fff; }
  .confidence .text { font-size: 14px; color: #1a1a1a; }
  .confidence .text b { display: block; margin-bottom: 4px; }
  .confidence .stats { margin-left: auto; font-size: 13px; color: #5f6368; white-space: nowrap; }
  /* Validation table */
  .val-row { display: grid; grid-template-columns: 90px 1fr 80px 110px; gap: 12px; padding: 10px 12px; border-bottom: 1px solid #f0f0f0; align-items: start; }
  .val-row:hover { background: #fafbfc; }
  .val-status { font-weight: 700; font-size: 12px; padding: 3px 8px; border-radius: 4px; text-align: center; }
  .val-status.PASS { background: #e6f4ea; color: #137333; }
  .val-status.WARN { background: #fef7e0; color: #9c5700; }
  .val-status.FAIL { background: #fce8e6; color: #c5221f; }
  .val-sev { font-size: 11px; font-weight: 600; padding: 2px 6px; border-radius: 3px; text-align: center; }
  .val-sev.HIGH { background: #ea4335; color: #fff; }
  .val-sev.MED { background: #f9ab00; color: #fff; }
  .val-sev.LOW { background: #dadce0; color: #3c4043; }
  .val-name { font-weight: 600; font-size: 13px; }
  .val-cat { font-size: 11px; color: #5f6368; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 2px; }
  .val-msg { font-size: 12px; color: #5f6368; margin-top: 2px; }
  .caveats { list-style: none; padding: 0; margin: 0; }
  .caveats li { padding: 6px 0; font-size: 13px; color: #1a1a1a; border-bottom: 1px solid #f0f0f0; }
  .caveats li:last-child { border-bottom: none; }
  .rec-caveat { background: #fff2cc; border-radius: 6px; padding: 8px 12px; font-size: 12px; margin-top: 6px; color: #6b5d00; }
  /* Goal table */
  .goal-good { color: #137333; font-weight: 600; }
  .goal-bad { color: #c5221f; font-weight: 600; }
  .goal-warn { color: #b06000; font-weight: 600; }
  tr.goal-row-good { background: #f1f8f4; }
  tr.goal-row-bad { background: #fcefee; }
  tr.goal-row-warn { background: #fef9e8; }
  tr.goal-row-good:hover, tr.goal-row-bad:hover, tr.goal-row-warn:hover { filter: brightness(0.97); }
</style>
</head>
<body>
<div class="container">
  <h1>Отчёт по сотрудникам</h1>
  <div class="subtitle">Период: {{period}} • Сгенерировано: {{generated}}</div>

  <div class="confidence {{conf_color}}">
    <div class="badge">{{conf_level}}</div>
    <div class="text">
      <b>Статус отчёта</b>
      {{conf_summary}}
    </div>
    <div class="stats">{{conf_stats}}</div>
  </div>

  <div class="cards">
    <div class="card">
      <div class="label">Общий доход</div>
      <div class="value">{{total_revenue}}</div>
      <div class="hint">за период наблюдения</div>
    </div>
    <div class="card">
      <div class="label">Устройств активно</div>
      <div class="value">{{total_devices}}</div>
      <div class="hint">по {{n_employees}} сотрудникам</div>
    </div>
    <div class="card">
      <div class="label">Средний доход</div>
      <div class="value">{{avg_per_dev}}</div>
      <div class="hint">на устройство</div>
    </div>
    <div class="card">
      <div class="label">Выполнение цели</div>
      <div class="value" style="color: {{goal_color}};">{{company_pct}}</div>
      <div class="hint">{{goal_hint}}</div>
    </div>
  </div>

  <section>
    <h2>Достижение цели · {{target_per_day}} ₽/устр/сутки</h2>
    <div style="font-size: 13px; color: #5f6368; margin-bottom: 16px;">
      План у каждого сотрудника считается индивидуально: <code>цель × устройств × период в сутках</code>.
      Сравнение по выполнению плана корректнее, чем по абсолютному доходу.
    </div>
    <table>
      <thead>
        <tr>
          <th>Сотрудник</th>
          <th class="num">Устр.</th>
          <th class="num">Период, сут</th>
          <th class="num">Факт, ₽/устр/сутки</th>
          <th class="num">Отклонение</th>
          <th class="num">План за период</th>
          <th class="num">Факт за период</th>
          <th class="num">Выполнение</th>
        </tr>
      </thead>
      <tbody>
        {{goal_rows}}
      </tbody>
      <tfoot>
        <tr style="background: #e8f0fe; font-weight: 600;">
          <td>ИТОГО по компании</td>
          <td class="num">{{total_devices_short}}</td>
          <td class="num">{{avg_period}}</td>
          <td class="num">{{company_fact_day}}</td>
          <td class="num">{{company_dev_day}}</td>
          <td class="num">{{total_plan}}</td>
          <td class="num">{{total_revenue}}</td>
          <td class="num"><b style="color: {{goal_color}};">{{company_pct}}</b></td>
        </tr>
      </tfoot>
    </table>
  </section>

  <section>
    <h2>Структура дохода по каналам</h2>
    <div class="legend">
      <span><i style="background:#1D9E75"></i>WhatsApp: {{ch_wa}}</span>
      <span><i style="background:#185FA5"></i>Regular: {{ch_reg}}</span>
      <span><i style="background:#BA7517"></i>Kakao: {{ch_kakao}}</span>
      <span><i style="background:#D85A30"></i>Click: {{ch_click}}</span>
      <span><i style="background:#888780"></i>Business: {{ch_biz}}</span>
    </div>
    <div class="channel-bar">
      <div style="background:#1D9E75; width:{{w_wa}}%" title="WhatsApp"></div>
      <div style="background:#185FA5; width:{{w_reg}}%" title="Regular"></div>
      <div style="background:#BA7517; width:{{w_kakao}}%" title="Kakao"></div>
      <div style="background:#D85A30; width:{{w_click}}%" title="Click"></div>
      <div style="background:#888780; width:{{w_biz}}%" title="Business"></div>
    </div>
  </section>

  <section>
    <h2>Эффективность по сервисам</h2>
    <div style="font-size: 13px; color: #5f6368; margin-bottom: 16px;">
      Конверсия кода = коды/старты · Бан = сгоревшие аккаунты/всего аккаунтов · Доставка = доставлено/отправлено · SMS/акк = отправлено на один аккаунт
    </div>
    <table>
      <thead>
        <tr>
          <th>Сервис</th>
          <th class="num">Старты</th>
          <th class="num">Коды</th>
          <th class="num">Аккаунты</th>
          <th class="num">Отправл.</th>
          <th class="num">Конверсия кода</th>
          <th class="num">Брак</th>
          <th class="num">Бан аккаунтов</th>
          <th class="num">Доставка</th>
          <th class="num">SMS/акк</th>
        </tr>
      </thead>
      <tbody>
        {{service_rows}}
      </tbody>
    </table>
  </section>

  <section>
    <h2>Рейтинг сотрудников</h2>
    <table>
      <thead>
        <tr>
          <th>#</th>
          <th>Сотрудник</th>
          <th class="num">Доход</th>
          <th class="num">Доля</th>
          <th class="num">Устройств</th>
          <th class="num">₽/устр</th>
          <th>Активность 24ч</th>
        </tr>
      </thead>
      <tbody>
        {{rating_rows}}
      </tbody>
    </table>
  </section>

  <section>
    <h2>Качество работы</h2>
    <table>
      <thead>
        <tr>
          <th>Сотрудник</th>
          <th class="num">Click брак</th>
          <th class="num">WA бан акк</th>
          <th class="num">WA доставка</th>
          <th class="num">Kakao конверсия</th>
          <th class="num">Мёртвых устр</th>
        </tr>
      </thead>
      <tbody>
        {{quality_rows}}
      </tbody>
    </table>
  </section>

  <section>
    <h2>Топ-15 устройств по доходу</h2>
    <table>
      <thead>
        <tr>
          <th>#</th>
          <th>Сотрудник</th>
          <th>API Key</th>
          <th>Модель</th>
          <th class="num">Доход</th>
        </tr>
      </thead>
      <tbody>
        {{top_dev_rows}}
      </tbody>
    </table>
  </section>

  <section>
    <h2>Проверки качества отчёта</h2>
    <div style="font-size: 13px; color: #5f6368; margin-bottom: 12px;">
      Автоматический чек-лист по мотивам <a href="https://github.com/anthropics/knowledge-work-plugins/blob/main/data/skills/validate-data/SKILL.md" target="_blank" style="color: #1967d2;">validate-data skill</a> от Anthropic. Проверяет корректность расчётов, разумность результатов и аналитические ловушки.
    </div>
    {{validation_rows}}
  </section>

  <section>
    <h2>Оговорки для презентации</h2>
    <div style="font-size: 13px; color: #5f6368; margin-bottom: 12px;">
      Эти моменты обязательно упомянуть при презентации руководству, чтобы данные были интерпретированы корректно.
    </div>
    <ul class="caveats">
      {{caveats_list}}
    </ul>
  </section>

  <section>
    <h2>Персональные рекомендации</h2>
    <div class="rec-grid">
      {{rec_blocks}}
    </div>
  </section>

  <section>
    <h2>Личные кабинеты сотрудников</h2>
    <div style="font-size: 13px; color: #5f6368; margin-bottom: 16px;">
      Каждый раздел можно раскрыть — внутри причины недотягивания, разбивка по платформам и список устройств (худшие сверху). Найди себя и разберись со своими устройствами.
    </div>
    {{employee_cards}}
  </section>

  <div class="footer">Автоматический отчёт • analyze_reports.py</div>
</div>
</body>
</html>
"""


def write_html(summary, top_dev, recs, period_info, checks, confidence, svc_eff, all_data, out_path,
               period_days: float = None):
    """Создаём HTML-отчёт."""
    def esc(x):
        return html.escape(str(x))

    total_devices_raw, total_revenue, total_device_days, avg_period_days, company_fact_day = company_goal_totals(summary)
    total_devices = int(total_devices_raw)
    avg_per_dev = total_revenue / total_devices if total_devices else 0

    ch_wa = summary['Доход_Wa'].sum()
    ch_reg = summary['Доход_Regular'].sum()
    ch_kakao = summary['Доход_Kakao'].sum()
    ch_click = summary['Доход_Click'].sum()
    ch_biz = summary['Доход_Business'].sum()

    def pct_of_total(x):
        return x / total_revenue * 100 if total_revenue else 0

    # Рейтинг
    rating_rows = []
    max_rev = summary['Итого_руб'].max() if len(summary) else 1
    for idx, row in summary.iterrows():
        share = pct_of_total(row['Итого_руб'])
        bar_w = (row['Итого_руб'] / max_rev * 100) if max_rev else 0
        act = row['Активность_24ч_pct']
        act_class = 'good' if act >= 30 else ('bad' if act < 15 else 'neutral')
        rating_rows.append(f"""
        <tr>
          <td>{idx+1}</td>
          <td><b>{esc(row['Сотрудник'])}</b></td>
          <td class="num bar" style="--w:{bar_w:.1f}%;">{fmt_rub(row['Итого_руб'])}</td>
          <td class="num">{share:.1f}%</td>
          <td class="num">{int(row['Устройств'])}</td>
          <td class="num">{row['Доход_на_устр']:.0f} ₽</td>
          <td><span class="{act_class}">{act:.0f}%</span></td>
        </tr>""")

    # Качество
    quality_rows = []
    for _, row in summary.iterrows():
        def chip_pct(value, good_threshold, bad_threshold, lower_is_better=False):
            if lower_is_better:
                if value <= good_threshold: cls = 'chip-green'
                elif value >= bad_threshold: cls = 'chip-red'
                else: cls = 'chip-yellow'
            else:
                if value >= good_threshold: cls = 'chip-green'
                elif value <= bad_threshold: cls = 'chip-red'
                else: cls = 'chip-yellow'
            return f'<span class="chip {cls}">{value:.1f}%</span>'

        dead_pct = row['Мертвых_pct']
        dead_chip_cls = 'chip-green' if dead_pct < 3 else ('chip-red' if dead_pct >= 7 else 'chip-yellow')

        quality_rows.append(f"""
        <tr>
          <td><b>{esc(row['Сотрудник'])}</b></td>
          <td class="num">{chip_pct(row['Click_bad_pct'], 30, 60, lower_is_better=True)}</td>
          <td class="num">{chip_pct(row['WA_ban_pct'], 8, 15, lower_is_better=True)}</td>
          <td class="num">{chip_pct(row['WA_delivery_pct'], 90, 80, lower_is_better=False)}</td>
          <td class="num">{chip_pct(row['Kakao_codes_pct'], 18, 10, lower_is_better=False)}</td>
          <td class="num"><span class="chip {dead_chip_cls}">{int(row['Мертвый'])} ({dead_pct:.1f}%)</span></td>
        </tr>""")

    # Топ-устройства
    top_dev_rows = []
    for idx, row in top_dev.head(15).iterrows():
        op = str(row.get('Operator', ''))
        model = op.split(' - ')[0] if op else ''
        top_dev_rows.append(f"""
        <tr>
          <td>{idx+1}</td>
          <td>{esc(row['_Сотрудник'])}</td>
          <td>{esc(row['API Key'])}</td>
          <td>{esc(model)}</td>
          <td class="num">{fmt_rub(row['Итого_руб'])}</td>
        </tr>""")

    # Рекомендации
    rec_blocks = []
    for rec in recs:
        caveat_html = ''
        if rec.get('Caveat'):
            caveat_html = f'<div class="rec-caveat">ⓘ Оговорка: {esc(rec["Caveat"])}</div>'
        rec_blocks.append(f"""
        <div class="rec">
          <div class="rec-head">
            <div class="rec-name">{esc(rec['Сотрудник'])}</div>
            <div class="rec-money">{fmt_rub(rec['Доход'])}</div>
          </div>
          <div class="rec-pluses">✓ {esc(rec['Плюсы'])}</div>
          <div class="rec-minuses">⚠ {esc(rec['Минусы'])}</div>
          {caveat_html}
        </div>""")

    leader_name = summary.iloc[0]['Сотрудник'] if len(summary) else '—'
    leader_share = f"{pct_of_total(summary.iloc[0]['Итого_руб']):.1f}%" if len(summary) else '—'

    # === ЦЕЛЕВЫЕ ПОКАЗАТЕЛИ ===
    total_plan = summary['План_за_период'].sum()
    company_pct = total_revenue / total_plan * 100 if total_plan else 0

    if company_pct >= 100:
        goal_color = '#137333'
        goal_hint = f'Цель достигнута: {company_fact_day:.0f} ₽/устр/сутки'
    elif company_pct >= 70:
        goal_color = '#b06000'
        goal_hint = f'Недовыполнение: {company_fact_day:.0f} ₽/устр/сутки'
    else:
        goal_color = '#c5221f'
        goal_hint = f'Существенный недобор: {company_fact_day:.0f} ₽/устр/сутки'

    # Сортировка для goal-таблицы — по выполнению %
    summary_by_goal = summary.sort_values('Выполнение_pct', ascending=False).reset_index(drop=True)
    goal_rows = []
    for _, row in summary_by_goal.iterrows():
        pct = row['Выполнение_pct']
        if pct >= 100:
            row_cls = 'goal-row-good'
            pct_cls = 'goal-good'
        elif pct >= 50:
            row_cls = 'goal-row-warn'
            pct_cls = 'goal-warn'
        else:
            row_cls = 'goal-row-bad'
            pct_cls = 'goal-bad'

        dev = row['Отклонение_сутки']
        dev_str = f'+{dev:.1f} ₽' if dev >= 0 else f'{dev:.1f} ₽'
        dev_cls = 'goal-good' if dev >= 0 else 'goal-bad'

        goal_rows.append(f"""
        <tr class="{row_cls}">
          <td><b>{esc(row['Сотрудник'])}</b></td>
          <td class="num">{int(row['Устройств'])}</td>
          <td class="num">{row['Период_суток']:.2f}</td>
          <td class="num">{row['Факт_руб_устр_сутки']:.1f} ₽</td>
          <td class="num {dev_cls}">{dev_str}</td>
          <td class="num">{fmt_rub(row['План_за_период'])}</td>
          <td class="num">{fmt_rub(row['Итого_руб'])}</td>
          <td class="num {pct_cls}">{pct:.0f}%</td>
        </tr>""")

    company_dev_day = company_fact_day - TARGET_PER_DEVICE_DAY
    company_dev_str = f'+{company_dev_day:.1f} ₽' if company_dev_day >= 0 else f'{company_dev_day:.1f} ₽'

    # Service efficiency rows
    service_rows = []
    if svc_eff is not None and not svc_eff.empty:
        for _, row in svc_eff.iterrows():
            conv = row['Конверсия_кода_pct']
            conv_cls = 'goal-good' if conv >= 20 else ('goal-bad' if conv < 5 and row['Старты'] > 0 else 'goal-warn')
            ban = row['Бан_аккаунтов_pct']
            ban_cls = 'goal-bad' if ban >= 15 else ('goal-good' if ban > 0 and ban < 8 else 'neutral')
            deliv = row['Доставка_pct']
            deliv_cls = 'goal-good' if deliv >= 90 else ('goal-bad' if deliv > 0 and deliv < 80 else 'neutral')
            conv_disp = f'{conv:.1f}%' if row['Старты'] > 0 else '—'
            ban_disp = f'{ban:.1f}%' if row['Аккаунты'] > 0 else '—'
            deliv_disp = f'{deliv:.1f}%' if row['Отправлено'] > 0 else '—'
            sms_disp = f'{row["SMS_на_аккаунт"]:.1f}' if row['Аккаунты'] > 0 else '—'
            brak_disp = f'{row["Брак_pct"]:.1f}%' if row['Старты'] > 0 else '—'
            service_rows.append(f"""
        <tr>
          <td><b>{esc(row['Сервис'])}</b></td>
          <td class="num">{row['Старты']:,}</td>
          <td class="num">{row['Коды']:,}</td>
          <td class="num">{row['Аккаунты']:,}</td>
          <td class="num">{row['Отправлено']:,}</td>
          <td class="num {conv_cls}">{conv_disp}</td>
          <td class="num">{brak_disp}</td>
          <td class="num {ban_cls}">{ban_disp}</td>
          <td class="num {deliv_cls}">{deliv_disp}</td>
          <td class="num">{sms_disp}</td>
        </tr>""".replace(',', ' '))

    # Персональные карточки сотрудников
    employee_cards = []
    dev_status_class = {
        'мёртвое (нет дохода, простой >2 дн)': 'dev-dead',
        'нулевой доход': 'dev-zero',
        'ниже половины цели': 'dev-low',
        'ниже цели': 'dev-low',
        'OK': 'dev-ok',
    }
    for emp_name in summary['Сотрудник']:
        detail = compute_employee_detail(all_data, emp_name, period_days=period_days)
        if detail is None:
            continue
        comp = detail['completion']
        badge_cls = 'green' if comp >= 100 else ('yellow' if comp >= 50 else 'red')
        reasons_html = ''.join(f'<li>{esc(rs)}</li>' for rs in detail['reasons'])

        api_rows = ''
        for _, ar in detail['by_api'].iterrows():
            api_rows += (f"<tr><td>{esc(ar['API Key'])}</td>"
                         f"<td class='num'>{int(ar['Устройств'])}</td>"
                         f"<td class='num'>{fmt_rub(ar['Доход'])}</td>"
                         f"<td class='num'>{ar['Факт_сутки']:.1f} ₽</td></tr>")

        svc_rows_emp = ''
        if detail['services'] is not None and not detail['services'].empty:
            for _, sr in detail['services'].iterrows():
                ban_disp = f"{sr['Бан_аккаунтов_pct']:.0f}%" if sr['Аккаунты'] > 0 else '—'
                deliv_disp = f"{sr['Доставка_pct']:.0f}%" if sr['Отправлено'] > 0 else '—'
                sms_disp = f"{sr['SMS_на_аккаунт']:.1f}" if sr['Аккаунты'] > 0 else '—'
                svc_rows_emp += (f"<tr><td>{esc(sr['Сервис'])}</td>"
                                 + f"<td class='num'>{sr['Старты']:,}</td>".replace(',', ' ')
                                 + f"<td class='num'>{sr['Коды']:,}</td>".replace(',', ' ')
                                 + f"<td class='num'>{ban_disp}</td>"
                                 f"<td class='num'>{deliv_disp}</td>"
                                 f"<td class='num'>{sms_disp}</td></tr>")

        dev_rows = ''
        devices_show = detail['devices'].head(25)
        for _, dr in devices_show.iterrows():
            cls = dev_status_class.get(dr['Статус'], '')
            dev_rows += (f"<tr class='{cls}'><td>{esc(str(dr['Ext ID']))}</td>"
                         f"<td>{esc(str(dr['API']))}</td>"
                         f"<td>{esc(str(dr['Модель']))}</td>"
                         f"<td class='num'>{fmt_rub(dr['Доход'])}</td>"
                         f"<td class='num'>{dr['Факт_сутки']:.1f} ₽</td>"
                         f"<td>{esc(dr['Статус'])}</td></tr>")
        more_note = ''
        if len(detail['devices']) > 25:
            more_note = f"<p style='font-size:12px;color:#9ca3af'>Показаны 25 худших из {len(detail['devices'])}. Полный список — в Excel-отчёте.</p>"

        svc_block = ''
        if svc_rows_emp:
            svc_block = (f"<h4>Эффективность сервисов</h4>"
                         f"<table class='mini'><thead><tr><th>Сервис</th><th class='num'>Старты</th>"
                         f"<th class='num'>Коды</th><th class='num'>Бан акк</th><th class='num'>Доставка</th>"
                         f"<th class='num'>SMS/акк</th></tr></thead><tbody>{svc_rows_emp}</tbody></table>")

        employee_cards.append(f"""
    <details class="emp-card">
      <summary>
        <span>{esc(emp_name)}</span>
        <span class="emp-badge {badge_cls}">{comp:.0f}% плана · {detail['fact_per_day']:.0f} ₽/устр/сутки</span>
      </summary>
      <div class="emp-body">
        <div class="emp-stats">
          <div>Доход за период<b>{fmt_rub(detail['total_revenue'])}</b></div>
          <div>Устройств<b>{detail['n_devices']}</b></div>
          <div>Период<b>{detail['period_days']:.1f} сут</b></div>
          <div>Факт/устр/сутки<b>{detail['fact_per_day']:.1f} ₽</b></div>
          <div>Цель<b>{TARGET_PER_DEVICE_DAY:.0f} ₽</b></div>
        </div>
        <h4>Почему не дотягиваем до плана</h4>
        <ul class="emp-reasons">{reasons_html}</ul>
        <h4>Платформы (API Key)</h4>
        <table class="mini">
          <thead><tr><th>API Key</th><th class="num">Устройств</th><th class="num">Доход</th><th class="num">₽/устр/сутки</th></tr></thead>
          <tbody>{api_rows}</tbody>
        </table>
        {svc_block}
        <h4>Устройства (худшие сверху — с них начинать)</h4>
        <table class="mini">
          <thead><tr><th>Ext ID</th><th>Платформа</th><th>Модель</th><th class="num">Доход</th><th class="num">₽/сутки</th><th>Статус</th></tr></thead>
          <tbody>{dev_rows}</tbody>
        </table>
        {more_note}
      </div>
    </details>""")

    # Validation rows (HTML)
    status_order = {'FAIL': 0, 'WARN': 1, 'PASS': 2}
    severity_order = {'HIGH': 0, 'MED': 1, 'LOW': 2}
    sorted_checks = sorted(
        checks,
        key=lambda c: (status_order.get(c['status'], 9), severity_order.get(c['severity'], 9))
    )
    val_rows = []
    for ch in sorted_checks:
        val_rows.append(f"""
        <div class="val-row">
          <div class="val-status {ch['status']}">{ch['status']}</div>
          <div>
            <div class="val-cat">{esc(ch['category'])}</div>
            <div class="val-name">{esc(ch['name'])}</div>
            <div class="val-msg">{esc(ch['message'])}</div>
          </div>
          <div class="val-sev {ch['severity']}">{ch['severity']}</div>
          <div></div>
        </div>""")

    # Caveats list
    caveats_html = []
    if confidence['caveats']:
        for cv in confidence['caveats'][:15]:
            caveats_html.append(f'<li>{esc(cv)}</li>')
    else:
        caveats_html.append('<li>Особых оговорок нет — все ключевые проверки пройдены.</li>')

    # Confidence color mapping
    conf_color_class = {'GREEN': 'green', 'YELLOW': 'yellow', 'RED': 'red'}.get(confidence['level_code'], 'yellow')

    replacements = {
        '{{period}}': esc(period_info),
        '{{generated}}': datetime.now().strftime('%Y-%m-%d %H:%M'),
        '{{total_revenue}}': fmt_rub(total_revenue),
        '{{total_devices}}': fmt_int(total_devices),
        '{{n_employees}}': str(len(summary)),
        '{{avg_per_dev}}': fmt_rub(avg_per_dev),
        '{{leader}}': esc(leader_name),
        '{{leader_share}}': leader_share,
        '{{ch_wa}}': fmt_rub(ch_wa),
        '{{ch_reg}}': fmt_rub(ch_reg),
        '{{ch_kakao}}': fmt_rub(ch_kakao),
        '{{ch_click}}': fmt_rub(ch_click),
        '{{ch_biz}}': fmt_rub(ch_biz),
        '{{w_wa}}': f"{pct_of_total(ch_wa):.2f}",
        '{{w_reg}}': f"{pct_of_total(ch_reg):.2f}",
        '{{w_kakao}}': f"{pct_of_total(ch_kakao):.2f}",
        '{{w_click}}': f"{pct_of_total(ch_click):.2f}",
        '{{w_biz}}': f"{pct_of_total(ch_biz):.2f}",
        '{{rating_rows}}': '\n'.join(rating_rows),
        '{{quality_rows}}': '\n'.join(quality_rows),
        '{{top_dev_rows}}': '\n'.join(top_dev_rows),
        '{{rec_blocks}}': '\n'.join(rec_blocks),
        '{{validation_rows}}': '\n'.join(val_rows),
        '{{caveats_list}}': '\n'.join(caveats_html),
        '{{conf_color}}': conf_color_class,
        '{{conf_level}}': esc(confidence['level']),
        '{{conf_summary}}': esc(confidence['summary']),
        '{{conf_stats}}': f"{confidence['n_pass']} PASS · {confidence['n_warn']} WARN · {confidence['n_fail']} FAIL",
        # Цель
        '{{target_per_day}}': f'{TARGET_PER_DEVICE_DAY:.0f}',
        '{{company_pct}}': f'{company_pct:.0f}%',
        '{{company_fact_day}}': f'{company_fact_day:.1f} ₽',
        '{{company_dev_day}}': company_dev_str,
        '{{total_plan}}': fmt_rub(total_plan),
        '{{avg_period}}': f'{avg_period_days:.2f}',
        '{{total_devices_short}}': fmt_int(total_devices),
        '{{goal_color}}': goal_color,
        '{{goal_hint}}': esc(goal_hint),
        '{{goal_rows}}': '\n'.join(goal_rows),
        '{{service_rows}}': '\n'.join(service_rows),
        '{{employee_cards}}': '\n'.join(employee_cards),
    }
    out = HTML_TEMPLATE
    for k, v in replacements.items():
        out = out.replace(k, v)

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(out)
    print(f"  HTML:  {out_path}")


# =============================================================
# 5. MAIN
# =============================================================

def main():
    in_folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('.')
    out_folder = Path(sys.argv[2]) if len(sys.argv) > 2 else in_folder

    in_folder = in_folder.resolve()
    out_folder = out_folder.resolve()
    out_folder.mkdir(parents=True, exist_ok=True)

    # В backend-конфигурации API и worker используют один общий volume /storage.
    config_candidates = [Path('/storage/employees.yml')]
    config_path = next((p for p in config_candidates if p.exists()), None)

    print(f"Папка с данными:  {in_folder}")
    print(f"Папка для отчёта: {out_folder}")
    if config_path:
        print(f"Конфиг:           {config_path}")
    print()

    print("Загрузка xlsx-файлов...")
    all_data, employees, mode_info = load_all_files(in_folder, config_path)
    print(f"  Загружено: {len(employees)} сотрудников, {len(all_data)} строк")

    # Предупреждение о неизвестных API Key
    if mode_info['unmatched_keys']:
        print(f"  ⚠ {len(mode_info['unmatched_keys'])} API Key из данных не найдены в конфиге:")
        for k in mode_info['unmatched_keys'][:20]:
            print(f"    - {k}")
        if len(mode_info['unmatched_keys']) > 20:
            print(f"    ... и ещё {len(mode_info['unmatched_keys']) - 20}")
        print("  Эти ключи попали в группу 'Неизвестно'. Поправь конфиг.")

    all_data = enrich(all_data)

    # Период наблюдения
    if 'Timestamp' in all_data.columns:
        ts = pd.to_datetime(all_data['Timestamp'], errors='coerce').dropna()
        if len(ts):
            period_info = f"{ts.min():%d.%m.%Y} — {ts.max():%d.%m.%Y}"
        else:
            period_info = 'не определён'
    else:
        period_info = 'не определён'

    print(f"  Период: {period_info}")

    print("Расчёт сводок...")
    summary = compute_summary(all_data)
    emp_totals = dict(zip(summary['Сотрудник'], summary['Итого_руб']))
    api_break = compute_api_breakdown(all_data, emp_totals)
    top_dev = compute_top_devices(all_data, n=30)
    model_stats = compute_model_stats(all_data)
    recs = build_recommendations(summary)
    svc_eff = compute_service_efficiency(all_data)
    n_services = len(svc_eff) if svc_eff is not None else 0
    print(f"  Активных сервисов: {n_services}")

    print("Валидация отчёта (sanity checks)...")
    checks = run_sanity_checks(all_data, summary, employees, mode_info)
    confidence = assess_confidence(checks, period_info, summary)
    print(f"  Уровень готовности: {confidence['level']}")
    print(f"  PASS: {confidence['n_pass']}  WARN: {confidence['n_warn']}  FAIL: {confidence['n_fail']}")

    excel_path = out_folder / 'Отчёт_по_сотрудникам.xlsx'
    html_path = out_folder / 'Отчёт_по_сотрудникам.html'
    methodology_path = out_folder / 'Methodology.md'

    print("Создание отчётов...")
    write_excel(summary, api_break, top_dev, model_stats, recs, period_info,
                checks, confidence, svc_eff, all_data, excel_path)
    write_html(summary, top_dev, recs, period_info, checks, confidence, svc_eff, all_data, html_path)

    # Methodology.md
    methodology_text = METHODOLOGY_TEMPLATE.format(
        generated=datetime.now().strftime('%Y-%m-%d %H:%M'),
        period=period_info,
    )
    with open(methodology_path, 'w', encoding='utf-8') as f:
        f.write(methodology_text)
    print(f"  Методология: {methodology_path}")

    print()
    print("Готово.")
    print(f"  Всего доход: {fmt_rub(summary['Итого_руб'].sum())}")
    print(f"  Сотрудников: {len(summary)}")
    print(f"  Устройств:   {int(summary['Устройств'].sum())}")
    print(f"  Статус:      {confidence['level']}")


if __name__ == '__main__':
    main()
